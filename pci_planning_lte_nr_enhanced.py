#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LTE/NR分离式PCI规划工具 - 支持mod30逻辑
主要特性：
1. 从待规划小区文件的LTE和NR工作表分别读取数据
2. NR小区使用mod30逻辑替代mod3逻辑
3. LTE和NR规划结果分别保存到不同文件
4. 优先确保最小PCI复用距离
"""

import pandas as pd
import numpy as np
import math
from typing import Dict, List, Tuple, Set, Optional, Union
from datetime import datetime
import time
import warnings
import zipfile
import os
import glob
from openpyxl import load_workbook

from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

class NetworkParameterUpdater:
    """
    现网工参更新工具类
    用于从现网工参压缩包中提取信息并更新全量工参文件
    支持模糊列名匹配和特殊处理逻辑
    """
    
    def __init__(self):
        self.full_param_dir = "全量工参"
        self.baseline_zip_pattern = "BaselineLab_*.zip"
        self.lte_online_params = None
        self.nr_online_params = None
        self.extracted_files = []
    
    def _get_latest_parameter_file(self, file_list):
        """从文件列表中选择时间戳最新的文件"""
        if not file_list:
            return None

        import re
        from datetime import datetime

        # 提取文件名中的时间戳并排序
        file_timestamps = []
        for file_path in file_list:
            file_name = os.path.basename(file_path)
            # 查找文件名中的14位时间戳 (YYYYMMDDHHMMSS)
            timestamp_match = re.search(r'(\d{14})', file_name)
            if timestamp_match:
                timestamp_str = timestamp_match.group(1)
                try:
                    # 尝试解析时间戳
                    timestamp = datetime.strptime(timestamp_str, '%Y%m%d%H%M%S')
                    file_timestamps.append((timestamp, file_path, f"14位时间戳: {timestamp_str}"))
                except ValueError:
                    # 如果不是14位时间戳，尝试查找8位日期
                    date_match = re.search(r'(\d{8})', file_name)
                    if date_match:
                        date_str = date_match.group(1)
                        try:
                            date = datetime.strptime(date_str, '%Y%m%d')
                            file_timestamps.append((date, file_path, f"8位日期: {date_str}"))
                        except ValueError:
                            # 如果都不是有效时间格式，使用文件修改时间
                            file_timestamps.append((datetime.fromtimestamp(os.path.getmtime(file_path)), file_path, f"文件修改时间"))
            else:
                # 如果没有找到时间戳，使用文件修改时间
                file_timestamps.append((datetime.fromtimestamp(os.path.getmtime(file_path)), file_path, f"文件修改时间"))

        # 按时间戳排序，选择最新的文件
        if file_timestamps:
            file_timestamps.sort(key=lambda x: x[0], reverse=True)
            latest_file = file_timestamps[0][1]
            timestamp_info = file_timestamps[0][2]
            print(f"选择最新文件: {os.path.basename(latest_file)} ({timestamp_info})")

            # 显示所有文件的时间戳信息（用于调试）
            print(f"找到 {len(file_timestamps)} 个候选文件:")
            for i, (timestamp, file_path, info) in enumerate(file_timestamps[:5]):  # 只显示前5个
                file_name = os.path.basename(file_path)
                print(f"  {i+1}. {file_name} - {info} ({timestamp.strftime('%Y-%m-%d %H:%M:%S')})")

            return latest_file

        return file_list[0] if file_list else None

    def update_network_parameters(self):
        """更新现网工参到全量工参"""
        try:
            # 查找全量工参文件
            full_param_files = glob.glob(f"{self.full_param_dir}/ProjectParameter_mongoose*.xlsx")
            if not full_param_files:
                print("❌ 未找到全量工参文件")
                return False

            # 选择最新的工参文件
            full_param_file = self._get_latest_parameter_file(full_param_files)
            if not full_param_file:
                print("❌ 无法确定最新的全量工参文件")
                return False

            print(f"找到最新的全量工参文件: {full_param_file}")
            
            # 查找现网工参压缩包
            baseline_zips = glob.glob(f"{self.full_param_dir}/{self.baseline_zip_pattern}")
            if not baseline_zips:
                print("❌ 未找到现网工参压缩包")
                return False
                
            baseline_zip = baseline_zips[0]
            print(f"正在提取现网工参压缩包: {baseline_zip}")
            
            # 提取压缩包中的工参文件
            lte_sdr_files, lte_itbbu_files, nr_files = self._extract_parameter_files(baseline_zip)
            
            if not any([lte_sdr_files, lte_itbbu_files, nr_files]):
                print("❌ 未找到有效的工参文件")
                return False
                
            # 加载现网工参数据
            lte_online_df = self._load_lte_online_data(lte_sdr_files, lte_itbbu_files)
            nr_online_df = self._load_nr_online_data(nr_files)
            
            if lte_online_df is None and nr_online_df is None:
                print("❌ 未能加载任何工参数据")
                return False
                
            # 加载全量工参数据
            print(f"开始更新全量工参文件: {full_param_file}")
            lte_full_df, nr_full_df = self._load_full_parameters(full_param_file)
            
            if lte_full_df is not None:
                print(f"全量LTE工参数: {len(lte_full_df)}")
            if nr_full_df is not None:
                print(f"全量NR工参数: {len(nr_full_df)}")
            
            # 更新工参数据
            updated = False
            if lte_online_df is not None and lte_full_df is not None:
                lte_full_df = self._update_lte_parameters(lte_full_df, lte_online_df)
                updated = True
                
            if nr_online_df is not None and nr_full_df is not None:
                nr_full_df = self._update_nr_parameters(nr_full_df, nr_online_df)
                updated = True
                
            if not updated:
                print("❌ 没有需要更新的工参数据")
                return False
                
            # 保存更新后的工参文件
            self._save_updated_parameters(full_param_file, lte_full_df, nr_full_df)
            
            print("✅ 现网工参更新完成")
            return True
            
        except Exception as e:
            print(f"❌ 更新全量工参失败: {e}")
            return False
    
    def _extract_parameter_files(self, zip_path):
        """从压缩包中提取工参文件"""
        lte_sdr_files = []
        lte_itbbu_files = []
        nr_files = []
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                # 创建临时目录存放解压文件
                extract_dir = "temp_extract"
                os.makedirs(extract_dir, exist_ok=True)
                
                # 只提取需要的文件
                for file_info in zip_ref.infolist():
                    filename = file_info.filename
                    # 处理路径分隔符问题，确保正确提取
                    clean_filename = filename.replace('\\', '/').split('/')[-1]
                    
                    if 'LTE_SDR_CellInfo' in clean_filename and clean_filename.endswith('.csv'):
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                        lte_sdr_files.append(extracted_path)
                    elif 'LTE_ITBBU_CellInfo' in clean_filename and clean_filename.endswith('.csv'):
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                        lte_itbbu_files.append(extracted_path)
                    elif 'NR_CellInfo' in clean_filename and clean_filename.endswith('.csv') and not clean_filename.endswith('.csvX'):
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                        nr_files.append(extracted_path)
                
                print(f"找到LTE_SDR文件: {len(lte_sdr_files)} 个")
                print(f"找到LTE_ITBBU文件: {len(lte_itbbu_files)} 个")
                print(f"找到NR文件: {len(nr_files)} 个")
                
                # 如果NR文件在根目录但路径处理有问题，重新检查
                if not nr_files:
                    for file_info in zip_ref.infolist():
                        filename = file_info.filename
                        clean_filename = filename.replace('\\', '/').split('/')[-1]
                        if 'NR_CellInfo' in clean_filename and clean_filename.endswith('.csv') and not clean_filename.endswith('.csvX'):
                            # 确保文件正确提取到临时目录
                            zip_ref.extract(file_info, extract_dir)
                            extracted_path = os.path.join(extract_dir, filename.replace('\\', '/'))
                            if os.path.exists(extracted_path):
                                nr_files.append(extracted_path)
                                print(f"重新提取NR文件: {clean_filename}")
                
        except Exception as e:
            print(f"提取压缩包文件失败: {e}")
            
        return lte_sdr_files, lte_itbbu_files, nr_files
    
    def _load_lte_online_data(self, lte_sdr_files, lte_itbbu_files):
        """加载LTE现网工参数据"""
        lte_dfs = []
        
        # 加载LTE_SDR文件
        for file in lte_sdr_files:
            try:
                df = pd.read_csv(file, encoding='utf-8')
                print(f"加载LTE_SDR文件: {os.path.basename(file)}, 记录数: {len(df)}")
                lte_dfs.append(df)
                # 清理临时文件
                os.remove(file)
            except Exception as e:
                print(f"加载LTE_SDR文件失败 {file}: {e}")
        
        # 加载LTE_ITBBU文件
        for file in lte_itbbu_files:
            try:
                df = pd.read_csv(file, encoding='utf-8')
                # 处理LTE_ITBBU文件的列名大小写问题 - 只标准化cellName相关列
                for col in df.columns:
                    if 'cellname' in col.lower():
                        # 将cellName相关列统一为'cellName'
                        df = df.rename(columns={col: 'cellName'})
                print(f"加载LTE_ITBBU文件: {os.path.basename(file)}, 记录数: {len(df)}")
                lte_dfs.append(df)
                # 清理临时文件
                os.remove(file)
            except Exception as e:
                print(f"加载LTE_ITBBU文件失败 {file}: {e}")
        
        # 清理临时目录
        temp_dir = "temp_extract"
        if os.path.exists(temp_dir) and os.path.isdir(temp_dir):
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        
        if not lte_dfs:
            return None
            
        # 合并所有LTE数据
        lte_online_df = pd.concat(lte_dfs, ignore_index=True)
        print(f"合并LTE现网工参总数: {len(lte_online_df)}")
        return lte_online_df
    
    def _load_nr_online_data(self, nr_files):
        """加载NR现网工参数据"""
        nr_dfs = []
        
        for file in nr_files:
            try:
                # 检查文件是否存在
                if not os.path.exists(file):
                    print(f"NR文件不存在: {file}")
                    # 尝试直接提取文件
                    # 动态查找现网工参压缩包
                    baseline_zips = glob.glob(f"{self.full_param_dir}/{self.baseline_zip_pattern}")
                    if baseline_zips:
                        zip_path = baseline_zips[0]
                        filename = os.path.basename(file)
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            zip_ref.extract(filename, "temp_extract")
                            extracted_file = os.path.join("temp_extract", filename)
                            if os.path.exists(extracted_file):
                                file = extracted_file
                                print(f"重新提取NR文件: {filename}")
                    
                if os.path.exists(file):
                    df = pd.read_csv(file, encoding='utf-8')
                    print(f"加载NR文件: {os.path.basename(file)}, 记录数: {len(df)}")
                    nr_dfs.append(df)
                    
                    # 显示NR文件的列名用于调试
                    print(f"NR文件列名: {list(df.columns)}")
                else:
                    print(f"NR文件仍然不存在: {file}")
                    
            except Exception as e:
                print(f"加载NR文件失败 {file}: {e}")
                # 尝试其他编码
                try:
                    df = pd.read_csv(file, encoding='gbk')
                    print(f"使用GBK编码加载NR文件成功: {os.path.basename(file)}, 记录数: {len(df)}")
                    nr_dfs.append(df)
                    print(f"NR文件列名: {list(df.columns)}")
                except Exception as e2:
                    print(f"使用GBK编码加载NR文件也失败 {file}: {e2}")
        
        # 清理临时文件
        for file in nr_files:
            try:
                if os.path.exists(file):
                    os.remove(file)
            except:
                pass
        
        # 清理临时目录
        temp_dir = "temp_extract"
        if os.path.exists(temp_dir) and os.path.isdir(temp_dir):
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        
        if not nr_dfs:
            return None
            
        # 合并所有NR数据
        nr_online_df = pd.concat(nr_dfs, ignore_index=True)
        print(f"合并NR现网工参总数: {len(nr_online_df)}")
        return nr_online_df
    
    def _load_full_parameters(self, file_path):
        """加载全量工参数据"""
        try:
            lte_full_df = None
            nr_full_df = None
            
            # 读取LTE工参表
            try:
                lte_full_df = pd.read_excel(file_path, sheet_name='LTE Project Parameters')
            except:
                print("未找到LTE工参表")
                
            # 读取NR工参表
            try:
                nr_full_df = pd.read_excel(file_path, sheet_name='NR Project Parameters')
            except:
                print("未找到NR工参表")
                
            return lte_full_df, nr_full_df
            
        except Exception as e:
            print(f"加载全量工参失败: {e}")
            return None, None
    
    def _apply_cell_formatting(self, file_path: str) -> None:
        """
        应用单元格格式设置 - 已移除颜色填充功能
        
        Args:
            file_path: Excel文件路径
        """
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            
            # 处理LTE工作表
            if 'LTE Project Parameters' in wb.sheetnames:
                lte_ws = wb['LTE Project Parameters']
                
                try:
                    # 安全地获取工作表的行列数
                    max_row = lte_ws.max_row
                    max_col = lte_ws.max_column
                    
                    # 检查工作表是否为空且有足够的行数
                    if max_row >= 4 and max_col > 0:
                        # 获取列名映射
                        lte_columns = {cell.value: cell.column for cell in lte_ws[1]}
                        
                        # 不再应用颜色填充
                        pass
                except Exception as e:
                    print(f"处理LTE工作表时出错: {e}")
            
            # 处理NR工作表
            if 'NR Project Parameters' in wb.sheetnames:
                nr_ws = wb['NR Project Parameters']
                
                try:
                    # 安全地获取工作表的行列数
                    max_row = nr_ws.max_row
                    max_col = nr_ws.max_column
                    
                    # 检查工作表是否为空且有足够的行数
                    if max_row >= 4 and max_col > 0:
                        # 获取列名映射
                        nr_columns = {cell.value: cell.column for cell in nr_ws[1]}
                        
                        # 不再应用颜色填充
                        pass
                except Exception as e:
                    print(f"处理NR工作表时出错: {e}")
            
            # 保存格式设置
            wb.save(file_path)
            print("✅ 单元格格式设置完成")
            
        except Exception as e:
            print(f"单元格格式设置失败: {e}")

    def _find_column_by_prefix(self, df: pd.DataFrame, prefix: str) -> str:
        """
        通过前缀模糊查找列名
        
        Args:
            df: 数据框
            prefix: 列名前缀
            
        Returns:
            匹配的列名，如果找不到返回None
        """
        for col in df.columns:
            if col.startswith(prefix):
                return col
        return None
    
    def _find_column_by_fuzzy_match(self, df: pd.DataFrame, keyword: str) -> str:
        """
        通过关键词模糊查找列名，支持部分匹配
        例如：列名为"子网ID SubNetwork ID long[0...128]"，只用"子网ID"即可匹配
        
        Args:
            df: 数据框
            keyword: 关键词
            
        Returns:
            匹配的列名，如果找不到返回None
        """
        # 首先尝试精确匹配
        if keyword in df.columns:
            return keyword
        
        # 然后尝试部分匹配 - 检查关键词是否包含在列名中
        for col in df.columns:
            if keyword in col:
                return col
        
        # 最后尝试前缀匹配作为备选
        return self._find_column_by_prefix(df, keyword)

        """
        更新LTE工参表
        
        Args:
            lte_full_df: 全量LTE工参
            lte_online_df: 现网LTE工参
            
        Returns:
            更新后的LTE工参
        """
        print("开始更新LTE工参表...")
        
        # 创建合并键
        # 使用模糊匹配查找列名
        enodeb_id_col = self._find_column_by_fuzzy_match(lte_full_df, 'eNodeB标识')
        cell_id_col = self._find_column_by_fuzzy_match(lte_full_df, '小区标识')
        print(f"LTE全量工参列名: {list(lte_full_df.columns)}")
        print(f"LTE现网工参列名: {list(lte_online_df.columns)}")
        print(f"模糊匹配结果 - eNodeB标识: {enodeb_id_col}, 小区标识: {cell_id_col}")
        lte_full_df['merge_key'] = lte_full_df[enodeb_id_col].astype(str) + "_" + lte_full_df[cell_id_col].astype(str)
        lte_online_df['merge_key'] = lte_online_df['eNBId'].astype(str) + "_" + lte_online_df['cellLocalId'].astype(str)
        
        # 字段映射 - 支持LTE_SDR和LTE_ITBBU文件的不同列名
        # 从现网工参字段 -> 全量工参字段
        field_mapping = {
            'SubNetwork': '子网ID',
            'ManagedElement': '管理网元ID',
            'eNBName': '基站名称',
            'eNBId': 'eNodeB标识',
            'cellName': '小区名称',  # LTE_SDR文件中的列名
            'CellName': '小区名称',  # LTE_ITBBU文件中的列名
            'cellLocalId': '小区标识',
            'tac': '跟踪区码',
            'pci': '物理小区识别码',
            'frequency': '下行链路的中心载频'
        }
        
        # 使用迭代方式更新，避免merge的性能问题
        updated_count = 0
        new_cells_count = 0
        
        # 创建全量工参的键值映射
        full_keys_set = set(lte_full_df['merge_key'])
        
        for online_idx, online_row in lte_online_df.iterrows():
            merge_key = online_row['merge_key']
            
            if merge_key in full_keys_set:
                # 更新现有记录
                full_mask = lte_full_df['merge_key'] == merge_key
                
                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(lte_full_df, full_field)
                        if target_col and pd.notna(online_row[online_field]):
                            lte_full_df.loc[full_mask, target_col] = online_row[online_field]
                            updated_count += 1
            else:
                # 添加新小区
                new_record = {}
                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(lte_full_df, full_field)
                        if target_col:
                            new_record[target_col] = online_row[online_field]
                        else:
                            # 当在全量工参中找不到对应列时，使用full_field作为列名并填充"未找到"
                            new_record[full_field] = "未找到"
                
                # 只添加有有效数据的记录
                if new_record:
                    lte_full_df = pd.concat([lte_full_df, pd.DataFrame([new_record])], ignore_index=True)
                    new_cells_count += 1
        
        print(f"LTE工参更新: 更新了 {updated_count} 个字段值")
        print(f"LTE工参新增: 添加了 {new_cells_count} 个新小区")
        
        # 移除临时列
        lte_full_df = lte_full_df.drop('merge_key', axis=1)
        
        # 填充默认值
        lte_full_df = self._fill_default_values(lte_full_df, 'LTE')
        
        print(f"LTE工参更新完成，总记录数: {len(lte_full_df)}")
        return lte_full_df
    
    def _fill_default_values(self, df: pd.DataFrame, network_type: str) -> pd.DataFrame:
        """
        填充默认值
        
        Args:
            df: 工参数据
            network_type: 网络类型 ('LTE' 或 'NR')
            
        Returns:
            填充默认值后的数据
        """
        if network_type == 'LTE':
            # 填充LTE默认值 - 强制填充
            df = self._fill_lte_defaults(df)
        elif network_type == 'NR':
            # 填充NR默认值 - 强制填充
            df = self._fill_nr_defaults(df)
        
        return df

    def _fill_lte_defaults(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        强制填充LTE默认值，跳过表头行（第1行和第2行）
        """
        # 模糊查找列名并强制填充，只对数据行（从第3行开始，索引2）进行填充
        
        for col in df.columns:
            # 第一分组 - 只对新增小区强制填入"电信中兴"，其他情况不操作
            if '第一分组' in col:
                # 只对值为空的单元格填入默认值，跳过表头行
                df.loc[2:, col] = df.loc[2:, col].fillna('电信中兴')
            # 系统制式 - 强制填入"1"，跳过表头行
            elif '系统制式' in col:
                df.loc[2:, col] = '1'
            # 移动国家码 - 强制填入"460"，跳过表头行
            elif '移动国家码' in col:
                df.loc[2:, col] = '460'
            # 移动网络码 - 默认填入"11"，跳过表头行
            elif '移动网络码' in col:
                df.loc[2:, col] = df.loc[2:, col].fillna('11')
            # 基站名称 - 保持为空时不处理
            elif '基站名称' in col:
                df[col] = df[col]  # 保持原值
            # 子网ID - 保持为空时不处理
            elif '子网ID' in col:
                df[col] = df[col]  # 保持原值
            # 管理网元ID - 保持为空时不处理
            elif '管理网元ID' in col:
                df[col] = df[col]  # 保持原值
        
        return df

    def _fill_nr_defaults(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        强制填充NR默认值，跳过表头行（第1行和第2行）
        """
        # 模糊查找列名并强制填充，只对数据行（从第3行开始，索引2）进行填充
        for col in df.columns:
            # 第一分组 - 只对新增小区强制填入"河源电信"，其他情况不操作
            if '第一分组' in col:
                # 只对值为空的单元格填入默认值，跳过表头行
                df.loc[2:, col] = df.loc[2:, col].fillna('河源电信')
            # 移动国家码 - 强制填入"460"，跳过表头行
            elif '移动国家码' in col:
                df.loc[2:, col] = '460'
            # gNodeBLength - 根据SSB频点判断
            elif 'gNodeBLength' in col:
                df = self._fill_gnodeb_length(df, col)
            # 移动网络码 - 从plmn字段提取最后两位
            elif '移动网络码' in col:
                df = self._fill_mnc_from_plmn(df, col)
            # 子网ID - 保持为空时不处理
            elif '子网ID' in col:
                df[col] = df[col]  # 保持原值
        
        return df

    def _fill_gnodeb_length(self, df: pd.DataFrame, length_col: str) -> pd.DataFrame:
        """
        根据SSB频点判断gNodeBLength值，跳过表头行（第1行和第2行）
        """
        # 查找SSB频点列
        ssb_col = None
        for col in df.columns:
            if 'SSB频点' in col or 'ssbFrequency' in col:
                ssb_col = col
                break
        
        if ssb_col is None:
            print("警告: 未找到SSB频点列，无法判断gNodeBLength")
            # 只对数据行（从第3行开始，索引2）填充默认值
            df.loc[2:, length_col] = df.loc[2:, length_col].fillna(5)
            return df
        
        # 根据SSB频点范围判断gNodeBLength
        def determine_gnodeb_length(ssb_freq):
            if pd.isna(ssb_freq):
                return "未找到"  # 当SSB频点为空时填充"未找到"
            
            try:
                ssb_freq = float(ssb_freq)
                if 870 <= ssb_freq <= 881:
                    return 5
                elif 2113 <= ssb_freq <= 2145:
                    return 1
                elif 3407 <= ssb_freq <= 3510:
                    return 78
                else:
                    return "未找到"  # 当SSB频点不在匹配范围时填充"未找到"
            except (ValueError, TypeError):
                return "未找到"  # 当SSB频点无法转换为数值时填充"未找到"
        
        # 只对数据行（从第3行开始，索引2）应用判断逻辑
        df.loc[2:, length_col] = df.loc[2:, ssb_col].apply(determine_gnodeb_length)
        return df

    def _fill_mnc_from_plmn(self, df: pd.DataFrame, mnc_col: str) -> pd.DataFrame:
        """
        从plmn字段提取移动网络码，跳过表头行（第1行和第2行）
        """
        # 查找plmn列
        plmn_col = None
        for col in df.columns:
            if 'plmn' in col.lower() or 'PLMN' in col:
                plmn_col = col
                break
        
        if plmn_col is None:
            print("警告: 未找到plmn列，无法提取移动网络码")
            # 不再填充默认值，保持原有值
            return df
        
        # 从plmn值提取移动网络码
        def extract_mnc_from_plmn(plmn_value):
            if pd.isna(plmn_value):
                return None  # 不设置默认值，保持为空
            
            plmn_str = str(plmn_value).strip()
            
            # 如果包含"-"
            if "-" in plmn_str:
                parts = plmn_str.split("-")
                if len(parts) >= 2:
                    right_part = parts[1]  # 右边字符（MNC）
                    
                    # 右边字符如果是"1"则改为"01"
                    if right_part == "1":
                        right_part = "01"
                    elif len(right_part) == 1 and right_part.isdigit():
                        # 其他单数字也补齐为2位
                        right_part = "0" + right_part
                    
                    return right_part
            
            # 如果不包含"-"，假设是46011这种格式
            if len(plmn_str) == 5 and plmn_str.startswith("460"):
                return plmn_str[3:]  # 提取后两位作为MNC
            
            # 保持原值
            return None
        
        # 只对数据行（从第3行开始，索引2）应用提取逻辑，且只更新空值
        mask = (df.index >= 2) & (df[mnc_col].isna())
        df.loc[mask, mnc_col] = df.loc[mask, plmn_col].apply(extract_mnc_from_plmn)
        return df
    
    def _update_lte_parameters(self, lte_full_df: pd.DataFrame, lte_online_df: pd.DataFrame) -> pd.DataFrame:
        """
        更新LTE工参表
        
        Args:
            lte_full_df: 全量LTE工参
            lte_online_df: 现网LTE工参
            
        Returns:
            更新后的LTE工参
        """
        print("开始更新LTE工参表...")
        
        # 创建合并键
        # 使用模糊匹配查找列名
        enodeb_id_col = self._find_column_by_fuzzy_match(lte_full_df, 'eNodeB标识')
        cell_id_col = self._find_column_by_fuzzy_match(lte_full_df, '小区标识')
        lte_full_df['merge_key'] = lte_full_df[enodeb_id_col].astype(str) + "_" + lte_full_df[cell_id_col].astype(str)
        lte_online_df['merge_key'] = lte_online_df['eNBId'].astype(str) + "_" + lte_online_df['cellLocalId'].astype(str)
        
        # 字段映射 - 支持LTE_SDR和LTE_ITBBU文件的不同列名
        field_mapping = {
            'SubNetwork': '子网ID',
            'ManagedElement': '管理网元ID',
            'eNBName': '基站名称',
            'eNBId': 'eNodeB标识',
            'cellName': '小区名称',  # LTE_SDR文件中的列名
            'CellName': '小区名称',  # LTE_ITBBU文件中的列名
            'cellLocalId': '小区标识',
            'tac': '跟踪区码',
            'pci': '物理小区识别码',
            'frequency': '下行链路的中心载频'
        }
        
        # 使用迭代方式更新，避免merge的性能问题
        updated_count = 0
        new_cells_count = 0
        
        # 创建全量工参的键值映射，跳过表头行（第1行和第2行，索引0和1）
        # 从第3行开始创建键值映射（索引从2开始）
        full_keys_set = set(lte_full_df['merge_key'][2:]) if len(lte_full_df) > 2 else set()
        
        for online_idx, online_row in lte_online_df.iterrows():
            merge_key = online_row['merge_key']
            
            if merge_key in full_keys_set:
                # 更新现有记录，跳过表头行（第1行和第2行，索引0和1）
                # 只从第3行开始查找匹配项（索引从2开始）
                matching_indices = lte_full_df[(lte_full_df['merge_key'] == merge_key) & (lte_full_df.index >= 2)].index
                
                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(lte_full_df, full_field)
                        if target_col and pd.notna(online_row[online_field]):
                            # 只更新数据行，跳过表头行
                            lte_full_df.loc[matching_indices, target_col] = online_row[online_field]
                            updated_count += len(matching_indices)
                        elif target_col is None:
                            # 当在全量工参中找不到对应列时，创建新列并填充"未找到"
                            # 只对数据行（从第3行开始，索引2）填充"未找到"
                            lte_full_df.loc[matching_indices, full_field] = "未找到"
                            updated_count += len(matching_indices)
            else:
                # 添加新小区
                new_record = {}
                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(lte_full_df, full_field)
                        if target_col:
                            new_record[target_col] = online_row[online_field]
                
                # 只添加有有效数据的记录
                if new_record:
                    lte_full_df = pd.concat([lte_full_df, pd.DataFrame([new_record])], ignore_index=True)
                    new_cells_count += 1
        
        print(f"LTE工参更新: 更新了 {updated_count} 个字段值")
        print(f"LTE工参新增: 添加了 {new_cells_count} 个新小区")
        
        # 移除临时列
        lte_full_df = lte_full_df.drop('merge_key', axis=1)
        
        # 填充默认值
        lte_full_df = self._fill_default_values(lte_full_df, 'LTE')
        
        print(f"LTE工参更新完成，总记录数: {len(lte_full_df)}")
        return lte_full_df
    
    def _update_nr_parameters(self, nr_full_df: pd.DataFrame, nr_online_df: pd.DataFrame) -> pd.DataFrame:
        """
        更新NR工参表
        
        Args:
            nr_full_df: 全量NR工参
            nr_online_df: 现网NR工参
            
        Returns:
            更新后的NR工参
        """
        print("开始更新NR工参表...")
        
        # 创建合并键 - 调整为全局匹配方式
        # 全量工参中的"移动国家码"&"移动网络码"&"gNodeB标识"&"小区标识"合并值
        mcc_col = self._find_column_by_fuzzy_match(nr_full_df, '移动国家码')
        mnc_col = self._find_column_by_fuzzy_match(nr_full_df, '移动网络码')
        gnodb_id_col = self._find_column_by_fuzzy_match(nr_full_df, 'gNodeB标识')
        cell_id_col = self._find_column_by_fuzzy_match(nr_full_df, '小区标识')
        
        # 处理plmn字段：删除"-"，并将值为"1"的移动网络码改为"01"
        def format_plmn(plmn):
            """
            处理plmn字段：删除"-"，并将值为"1"的移动网络码改为"01"
            """
            if pd.isna(plmn):
                return None  # 不设置默认值，保持为空
            
            plmn_str = str(plmn).strip()
            
            # 如果包含"-"
            if "-" in plmn_str:
                parts = plmn_str.split("-")
                if len(parts) >= 2:
                    left_part = parts[0]  # 左边字符（MCC）
                    right_part = parts[1]  # 右边字符（MNC）
                    
                    # 右边字符如果是"1"则改为"01"
                    if right_part == "1":
                        right_part = "01"
                    elif len(right_part) == 1 and right_part.isdigit():
                        # 其他单数字也补齐为2位
                        right_part = "0" + right_part
                    
                    return left_part + right_part
            
            # 如果不包含"-"，假设是46011这种格式
            if len(plmn_str) == 5 and plmn_str.startswith("460"):
                # 检查是否需要将"1"改为"01"
                if plmn_str.endswith("1") and not plmn_str.endswith("01") and not plmn_str.endswith("11") and not plmn_str.endswith("21") and not plmn_str.endswith("31") and not plmn_str.endswith("41") and not plmn_str.endswith("51") and not plmn_str.endswith("61") and not plmn_str.endswith("71") and not plmn_str.endswith("81") and not plmn_str.endswith("91"):
                    return plmn_str[:-1] + "01"
                return plmn_str
            
            # 保持原值
            return None
        
        # 确保移动网络码为2位字符串，不足用0补齐
        # 修复：处理"460-XX"格式，只保留"XX"部分
        def format_mnc(mnc):
            if pd.notna(mnc) and str(mnc).strip():
                mnc_str = str(mnc).strip()
                # 如果是"460-XX"格式，只保留"XX"部分
                if mnc_str.startswith("460-"):
                    mnc_part = mnc_str[4:]  # 去掉"460-"前缀
                    return mnc_part.zfill(2) if mnc_part.isdigit() else mnc_part
                # 其他情况保持原有逻辑
                return mnc_str.zfill(2) if mnc_str.isdigit() else mnc_str
            return None  # 不设置默认值，保持为空
        
        # 处理全量工参的移动网络码字段，应用format_mnc函数
        nr_full_df[mnc_col] = nr_full_df[mnc_col].apply(format_mnc)
        
        # 创建全量工参合并键
        nr_full_df[mcc_col] = nr_full_df[mcc_col].fillna('460')  # 填充缺失的移动国家码
        # 不再强制格式化移动网络码，保持原有值，只在匹配时更新
        nr_full_df['merge_key'] = (nr_full_df[mcc_col].astype(str) + 
                                  nr_full_df[mnc_col].astype(str) + 
                                  nr_full_df[gnodb_id_col].astype(str) + 
                                  nr_full_df[cell_id_col].astype(str))
        
        # 现网工参中的"plmn"&"gNBId"&"cellLocalId"合并值（使用新的处理方法）
        nr_online_df['processed_plmn'] = nr_online_df['plmn'].apply(format_plmn)
        nr_online_df['merge_key'] = (nr_online_df['processed_plmn'].astype(str) + 
                                    nr_online_df['gNBId'].astype(str) + 
                                    nr_online_df['cellLocalId'].astype(str))
        
        print(f"NR全量工参列名: {list(nr_full_df.columns)}")
        print(f"NR现网工参列名: {list(nr_online_df.columns)}")
        print(f"模糊匹配结果 - 移动国家码: {mcc_col}, 移动网络码: {mnc_col}, gNodeB标识: {gnodb_id_col}, 小区标识: {cell_id_col}")
        
        # 字段映射 - 从现网工参字段 -> 全量工参字段
        field_mapping = {
            'SubNetwork': '子网ID',
            'plmn': '移动网络码',
            'gNBName': '基站名称',
            'gNBId': 'gNodeB标识',
            'CellName': '小区名称',
            'cellLocalId': '小区标识',
            'ssbFrequency': '填写SSB频点',
            'pci': '物理小区识别码',
            'gNodeBLength': 'gNodeBLength'
        }
        
        # 使用迭代方式更新，避免merge的性能问题
        updated_count = 0
        new_cells_count = 0
        
        # 创建全量工参的键值映射，跳过表头行（第1行和第2行，索引0和1）
        # 从第3行开始创建键值映射（索引从2开始）
        full_keys_set = set(nr_full_df['merge_key'][2:]) if len(nr_full_df) > 2 else set()
        
        # 更新现有记录
        updated_count = 0
        for online_idx, online_row in nr_online_df.iterrows():
            merge_key = online_row['merge_key']
            
            if merge_key in full_keys_set:
                # 更新现有记录，跳过表头行（第1行和第2行，索引0和1）
                # 只从第3行开始查找匹配项（索引从2开始）
                matching_indices = nr_full_df[(nr_full_df['merge_key'] == merge_key) & (nr_full_df.index >= 2)].index
                
                # 只更新第一条匹配的记录，避免重复更新
                if len(matching_indices) > 0:
                    first_matching_index = matching_indices[0]
                    
                    for online_field, full_field in field_mapping.items():
                        if online_field in online_row:
                            # 使用模糊匹配查找全量工参中的目标列
                            target_col = self._find_column_by_fuzzy_match(nr_full_df, full_field)
                            if target_col and pd.notna(online_row[online_field]):
                                # 特殊处理plmn字段，应用format_mnc函数提取和格式化MNC
                                if online_field == 'plmn':
                                    nr_full_df.loc[first_matching_index, target_col] = format_mnc(online_row[online_field])
                                else:
                                    # 只更新数据行，跳过表头行
                                    nr_full_df.loc[first_matching_index, target_col] = online_row[online_field]
                                updated_count += 1
                    
                    # 更新full_keys_set中的键值，确保后续匹配正确
                    # 重新计算该记录的合并键
                    new_merge_key = (str(nr_full_df.loc[first_matching_index, mcc_col]) + 
                                   str(nr_full_df.loc[first_matching_index, mnc_col]) + 
                                   str(nr_full_df.loc[first_matching_index, gnodb_id_col]) + 
                                   str(nr_full_df.loc[first_matching_index, cell_id_col]))
                    # 从集合中移除旧键值并添加新键值
                    full_keys_set.discard(merge_key)
                    full_keys_set.add(new_merge_key)
            else:
                # 添加新小区
                new_record = {}
                for online_field, full_field in field_mapping.items():
                    if online_field in online_row:
                        # 使用模糊匹配查找全量工参中的目标列
                        target_col = self._find_column_by_fuzzy_match(nr_full_df, full_field)
                        if target_col:
                            # 特殊处理plmn字段，应用format_mnc函数提取和格式化MNC
                            if online_field == 'plmn':
                                new_record[target_col] = format_mnc(online_row[online_field])
                            else:
                                new_record[target_col] = online_row[online_field]
                
                # 只添加有有效数据的记录，并确保不重复
                if new_record:
                    # 创建新记录的合并键用于去重检查
                    new_gnodeb_id = new_record.get(gnodb_id_col, '')
                    new_cell_id = new_record.get(cell_id_col, '')
                    new_mcc = new_record.get(mcc_col, '460')
                    new_mnc = new_record.get(mnc_col, '')
                    
                    # 生成新记录的合并键
                    new_merge_key = str(new_mcc) + str(new_mnc) + str(new_gnodeb_id) + str(new_cell_id)
                    
                    # 检查是否已存在相同的记录（跳过表头行）
                    if new_merge_key not in full_keys_set:
                        nr_full_df = pd.concat([nr_full_df, pd.DataFrame([new_record])], ignore_index=True)
                        new_cells_count += 1
                        # 更新键值集合
                        full_keys_set.add(new_merge_key)
                    # 如果已存在，则跳过添加
        
        print(f"NR工参更新: 更新了 {updated_count} 个字段值")
        print(f"NR工参新增: 添加了 {new_cells_count} 个新小区")
        
        # 对NR工参进行去重处理，基于移动国家码、移动网络码、gNodeB标识和小区标识的组合
        print(f"去重前记录数: {len(nr_full_df)}")
        # 跳过表头行（第1行和第2行，索引0和1）
        header_df = nr_full_df.iloc[:2] if len(nr_full_df) >= 2 else nr_full_df
        data_df = nr_full_df.iloc[2:] if len(nr_full_df) > 2 else pd.DataFrame()
        
        if not data_df.empty:
            # 基于移动国家码、移动网络码、gNodeB标识和小区标识的组合进行去重，保留第一条记录
            data_df = data_df.drop_duplicates(subset=[mcc_col, mnc_col, gnodb_id_col, cell_id_col], keep='first')
            # 重新合并表头和数据
            nr_full_df = pd.concat([header_df, data_df], ignore_index=True)
        
        print(f"去重后记录数: {len(nr_full_df)}")
        
        # 移除临时列
        nr_full_df = nr_full_df.drop('merge_key', axis=1)
        if 'processed_plmn' in nr_online_df.columns:
            nr_online_df = nr_online_df.drop('processed_plmn', axis=1)
        
        # 填充默认值
        nr_full_df = self._fill_default_values(nr_full_df, 'NR')
        
        print(f"NR工参更新完成，总记录数: {len(nr_full_df)}")
        return nr_full_df
    
    def _save_updated_parameters(self, file_path, lte_full_df, nr_full_df):
        """保存更新后的工参文件"""
        try:
            # 生成新的文件名，保持原有结构，只更新日期部分
            import os
            import re
            from datetime import datetime

            # 获取当前时间戳 (YYYYMMDDHHMMSS格式) 确保唯一性
            current_timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

            # 提取原文件名中的信息
            base_name = os.path.basename(file_path)
            dir_path = os.path.dirname(file_path)

            # 检查文件名格式，确定时间戳的位置和长度
            # 模式1: ProjectParameter_mongoose河源电联20250928183756.xlsx (14位时间戳在末尾)
            # 模式2: ProjectParameter_mongoose20250923河源电联.xlsx (8位日期在中间)

            name_parts = os.path.splitext(base_name)
            file_name_without_ext = name_parts[0]
            file_ext = name_parts[1]

            # 尝试匹配末尾的14位时间戳 (YYYYMMDDHHMMSS)
            timestamp_match = re.search(r'(\d{14})$', file_name_without_ext)
            if timestamp_match:
                # 替换末尾的14位时间戳
                original_timestamp = timestamp_match.group(1)
                new_filename = file_name_without_ext[:-14] + current_timestamp + file_ext
                print(f"检测到末尾时间戳 {original_timestamp}，替换为 {current_timestamp}")
            else:
                # 尝试匹配末尾的8位日期 (YYYYMMDD)
                date_match = re.search(r'(\d{8})$', file_name_without_ext)
                if date_match:
                    # 替换末尾的8位日期为当前日期部分
                    original_date = date_match.group(1)
                    current_date = current_timestamp[:8]  # 只取前8位日期部分
                    new_filename = file_name_without_ext[:-8] + current_date + file_ext
                    print(f"检测到末尾日期 {original_date}，替换为 {current_date}")
                else:
                    # 如果没有找到时间戳模式，查找文件名中的8位或14位数字
                    all_numbers = re.findall(r'\d+', file_name_without_ext)
                    if all_numbers:
                        # 找到最大的数字序列并替换
                        max_number = max(all_numbers, key=len)
                        if len(max_number) == 14:
                            # 替换14位时间戳
                            new_filename = file_name_without_ext.replace(max_number, current_timestamp)
                        elif len(max_number) == 8:
                            # 替换8位日期
                            current_date = current_timestamp[:8]
                            new_filename = file_name_without_ext.replace(max_number, current_date)
                        else:
                            # 其他长度，在末尾添加时间戳
                            new_filename = file_name_without_ext + current_timestamp + file_ext
                    else:
                        # 如果没有数字，在末尾添加时间戳
                        new_filename = file_name_without_ext + current_timestamp + file_ext

            # 构建新的完整文件路径
            new_file_path = os.path.join(dir_path, new_filename)

            print(f"正在保存更新后的工参文件: {new_file_path}")
            print(f"原文件名: {base_name} -> 新文件名: {new_filename}")

            with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                if lte_full_df is not None:
                    lte_full_df.to_excel(writer, sheet_name='LTE Project Parameters', index=False)
                if nr_full_df is not None:
                    nr_full_df.to_excel(writer, sheet_name='NR Project Parameters', index=False)
            print("✅ 工参文件保存成功")
        except Exception as e:
            print(f"保存工参文件失败: {e}")
            # 如果追加模式失败，尝试创建新文件
            try:
                with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                    if lte_full_df is not None:
                        lte_full_df.to_excel(writer, sheet_name='LTE Project Parameters', index=False)
                    if nr_full_df is not None:
                        nr_full_df.to_excel(writer, sheet_name='NR Project Parameters', index=False)
                print("✅ 工参文件保存成功（新文件）")
            except Exception as e2:
                print(f"创建新工参文件失败: {e2}")


class LTENRPCIPlanner:
    def __init__(self, reuse_distance_km: float = 3.0, lte_inherit_mod3: bool = False,
                 nr_inherit_mod30: bool = False, network_type: str = "LTE", params_file: str = None):
        """
        初始化LTE/NR分离式PCI规划工具

        Args:
            reuse_distance_km: 最小PCI复用距离（公里）
            lte_inherit_mod3: LTE小区是否继承原PCI的模3值
            nr_inherit_mod30: NR小区是否继承原PCI的模30值
            network_type: 当前处理的网络类型 ("LTE" 或 "NR")
            params_file: 使用的参数文件路径
        """
        self.reuse_distance_km = reuse_distance_km
        self.lte_inherit_mod3 = lte_inherit_mod3
        self.nr_inherit_mod30 = nr_inherit_mod30
        self.network_type = network_type
        self.params_file = params_file  # 存储参数文件路径
        
        # PCI范围根据网络类型设定
        if network_type == "LTE":
            self.pci_range = list(range(0, 504))  # LTE PCI范围 0-503
            self.inherit_mod = lte_inherit_mod3
            self.mod_value = 3  # LTE使用mod3
            self.dual_mod_requirement = False  # LTE只需要模3约束
        else:  # NR
            self.pci_range = list(range(0, 1008))  # NR PCI范围 0-1007
            self.inherit_mod = nr_inherit_mod30
            self.mod_value = 30  # NR使用mod30
            self.dual_mod_requirement = True   # NR需要同时满足模3和模30约束
        
        # 数据存储
        self.cells_to_plan = None
        self.target_cells = None  # 目标网络类型的小区数据
        self.all_cells_combined = None  # 合并的所有小区数据（用于干扰分析）
        
        # 性能优化缓存
        self.distance_cache = {}  # 距离计算结果缓存
        self.pci_validity_cache = {}  # PCI验证结果缓存
        self.same_site_cache = {}  # 同站点信息缓存
        
        # 失败原因统计
        self.failure_reasons = {
            'cell_not_found': [],
            'no_location': [],
            'reuse_distance_violation': [],
            'no_compliant_pci': [],
            'fallback_assignments': [],
            'same_site_mod_conflicts': []  # 同站点模值冲突统计
        }
        
        print(f"初始化{network_type}网络PCI规划工具")
        print(f"同频PCI最小复用距离: {reuse_distance_km}km (优先级最高)")
        print(f"PCI分配策略 (按优先级):")
        print(f"  1. 最小复用距离 (最高优先级) - 确保同频同PCI小区间距离≥{reuse_distance_km}km")
        print(f"  2. 同站点模{self.mod_value}冲突避免 (第二优先级) - 避免同基站小区模值相同")
        print(f"  3. PCI分布均衡性 (第三优先级) - 优选复用距离接近阈值的PCI")
        
        if network_type == "LTE":
            if lte_inherit_mod3:
                print(f"模3继承: 是 - 严格按照原PCI的mod3值分配")
            else:
                print(f"模3继承: 否 - 自由规划，不考虑mod3值")
        else:
            if nr_inherit_mod30:
                print(f"模30继承: 是 - 严格按照原PCI的mod30值分配")
            else:
                print(f"模30继承: 否 - 自由规划，不考虑mod30值")
                
        print(f"PCI范围: 0-{max(self.pci_range)}")
        print(f"核心规则: 只有同频(earfcnDl相同)且PCI相同的小区才需要遵循最小复用距离")
    
    def generate_timestamp_suffix(self) -> str:
        """
        生成时间后缀，格式：YYYYMMDD_HHMMSS
        """
        now = datetime.now()
        return now.strftime("%Y%m%d_%H%M%S")
    
    def convert_to_numeric(self, df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
        """
        转换指定列为数值格式，处理文本格式数据
        """
        df_copy = df.copy()
        
        for col in columns:
            if col in df_copy.columns:
                # 转换为字符串，处理各种数据类型
                df_copy[col] = df_copy[col].astype(str)
                
                # 移除空格和特殊字符
                df_copy[col] = df_copy[col].str.replace(' ', '').str.replace(',', '')
                
                # 处理 'nan', 'None', 空字符串等
                df_copy[col] = df_copy[col].replace(['nan', 'None', '', 'null'], np.nan)
                
                # 转换为数值
                df_copy[col] = pd.to_numeric(df_copy[col], errors='coerce')
        
        return df_copy
    
    def load_data(self, cells_file: str, params_file: str):
        """
        加载数据文件 - 支持分离式LTE/NR工作表
        """
        print(f"\\n加载数据文件...")
        print(f"规划小区文件: {cells_file}")
        print(f"参数文件: {params_file}")
        print(f"使用的全量工参文件: {os.path.basename(params_file)}")  # 明确显示使用的参数文件
        
        try:
            # 加载规划小区数据 - 从对应的工作表读取
            worksheet_name = self.network_type  # "LTE" 或 "NR"
            print(f"正在从'{worksheet_name}'工作表加载待规划小区...")
            
            self.cells_to_plan = pd.read_excel(cells_file, sheet_name=worksheet_name)
            print(f"规划小区数据加载成功: {len(self.cells_to_plan)} 行")
            
            # 列名标准化检查
            if self.network_type == "NR" and 'gNodeBID' in self.cells_to_plan.columns:
                # NR网络应使用gNodeBID列，不需要映射
                print("NR网络使用gNodeBID列作为基站标识")
            elif self.network_type == "LTE" and 'eNodeBID' in self.cells_to_plan.columns:
                print("LTE网络使用eNodeBID列作为基站标识")
            
            # 加载参数文件，根据网络类型选择工作表
            if self.network_type == "LTE":
                params_worksheet = "LTE Project Parameters"
            else:
                params_worksheet = "NR Project Parameters"
            
            print(f"正在加载{params_worksheet}工作表...")
            
            # 读取目标网络类型的数据
            target_raw = pd.read_excel(params_file, sheet_name=params_worksheet)
            self.target_cells = self.preprocess_target_cells(target_raw)
            
            # 只使用当前网络类型的数据进行干扰分析（不跨网络类型）
            self.all_cells_combined = self.target_cells.copy()
            print(f"加载{self.network_type}小区数据: {len(self.all_cells_combined)} 个")
                
        except Exception as e:
            print(f"错误: 无法加载{self.network_type}工作表: {e}")
            raise
    
    def preprocess_target_cells(self, raw_df: pd.DataFrame) -> pd.DataFrame:
        """
        预处理目标网络类型的小区数据
        """
        if raw_df.empty:
            return pd.DataFrame()
        
        processed = pd.DataFrame()
        
        if self.network_type == "LTE":
            # LTE列映射
            processed['enodeb_id'] = raw_df['eNodeB标识\neNodeB ID\nlong:[0..1048575]']
            processed['cell_id'] = raw_df['小区标识\ncellLocalId\ninteger:[0~2147483647]']
            processed['cell_name'] = raw_df['小区名称\nuserLabel\nstring[0..128]']
            processed['pci'] = raw_df['物理小区识别码\nPCI\nlong:[0..503]']
            processed['lat'] = raw_df['小区纬度\neNodeB Latitude\ndouble:[-90..90]']
            processed['lon'] = raw_df['小区经度\neNodeB Longitude double:[-180..180]']
            processed['earfcn_dl'] = raw_df['下行链路的中心载频\nearfcnDl\ndouble Step：0.1 \nUnite：MHz']
        else:
            # NR列映射  
            processed['enodeb_id'] = raw_df['gNodeB标识\ngNodeB ID\nLong:[0..1048575]']
            processed['cell_id'] = raw_df['小区标识\ncellLocalId\nInteger:[0~2147483647]']
            processed['cell_name'] = raw_df['小区名称\nCELL NAME\nString[0..128]']
            processed['pci'] = raw_df['物理小区识别码\nPCI\nLong:[0..1007]']
            processed['lat'] = raw_df['小区纬度\nCell  Latitude\nDouble:[-90..90]']
            processed['lon'] = raw_df['小区经度\nCell  Longitude Double:[-180..180]']
            processed['earfcn_dl'] = raw_df['填写SSB频点\nSSB Frequency\nDouble Step：0.01 \nUnite：MHz']
        
        processed['cell_type'] = self.network_type
        
        # 转换所有数值列（cell_name保持字符串格式）
        print(f"转换{self.network_type}小区数值格式...")
        numeric_cols = ['enodeb_id', 'cell_id', 'pci', 'lat', 'lon', 'earfcn_dl']
        processed = self.convert_to_numeric(processed, numeric_cols)
        
        # 过滤掉缺少关键信息的小区
        before_count = len(processed)
        
        # 记录被移除的小区信息
        invalid_cells = processed[processed['enodeb_id'].isna() | processed['cell_id'].isna()]
        
        processed = processed.dropna(subset=['enodeb_id', 'cell_id'])
        after_count = len(processed)
        
        if before_count != after_count:
            print(f"{self.network_type}数据清理: 移除了 {before_count - after_count} 个缺少关键信息的记录")
            if not invalid_cells.empty:
                print("被移除的小区信息:")
                for idx, cell in invalid_cells.iterrows():
                    enodeb_id = cell.get('enodeb_id', '缺失')
                    cell_id = cell.get('cell_id', '缺失')
                    cell_name = cell.get('cell_name', '未知')
                    print(f"  - 基站ID: {enodeb_id}, 小区ID: {cell_id}, 小区名称: {cell_name}")
        
        print(f"{self.network_type}小区预处理完成，有效小区数量: {len(processed)}")
        return processed
    
    
    def calculate_distance_vectorized(self, lat1, lon1, lat2_array, lon2_array):
        """
        向量化距离计算（带缓存优化）
        """
        # 生成缓存键
        cache_key = (round(lat1, 6), round(lon1, 6), 
                    tuple(np.round(lat2_array, 6)), tuple(np.round(lon2_array, 6)))
        
        if cache_key in self.distance_cache:
            return self.distance_cache[cache_key]
        
        # 转换为弧度
        lat1_rad = math.radians(lat1)
        lon1_rad = math.radians(lon1)
        lat2_rad = np.radians(lat2_array)
        lon2_rad = np.radians(lon2_array)
        
        # Haversine公式
        dlat = lat2_rad - lat1_rad
        dlon = lon2_rad - lon1_rad
        a = np.sin(dlat/2)**2 + math.cos(lat1_rad) * np.cos(lat2_rad) * np.sin(dlon/2)**2
        c = 2 * np.arcsin(np.sqrt(a))
        
        result = c * 6371  # 地球半径（公里）
        self.distance_cache[cache_key] = result
        return result
    
    def get_cell_info(self, enodeb_id: int, cell_id: int) -> Tuple[Optional[pd.Series], str]:
        """
        获取小区详细信息
        
        Returns:
            (cell_info, status)
        """
        # 在所有小区中查找（包含完整的位置信息）
        mask = (self.all_cells_combined['enodeb_id'] == enodeb_id) & (self.all_cells_combined['cell_id'] == cell_id)
        matches = self.all_cells_combined[mask]
        
        if matches.empty:
            return None, 'cell_not_found'
        
        cell_info = matches.iloc[0]
        
        # 检查位置信息
        if pd.isna(cell_info['lat']) or pd.isna(cell_info['lon']):
            return cell_info, 'no_location'
        
        return cell_info, 'success'
    
    def validate_pci_reuse_distance(self, candidate_pci: int, target_lat: float, target_lon: float, 
                                   target_earfcn: float, exclude_enodeb: int = None, exclude_cell: int = None) -> Tuple[bool, float]:
        """
        验证候选PCI是否满足最小复用距离要求（带缓存优化）
        """
        if self.all_cells_combined is None or self.all_cells_combined.empty:
            return True, float('inf')
        
        # 生成缓存键 - 关键修复：包含复用距离参数
        cache_key = (candidate_pci, round(target_lat, 6), round(target_lon, 6), 
                    round(target_earfcn, 2), exclude_enodeb, exclude_cell, 
                    round(self.reuse_distance_km, 2))
        
        if cache_key in self.pci_validity_cache:
            return self.pci_validity_cache[cache_key]
        
        # 只查找同频同PCI的小区，并且只考虑同网络类型
        same_freq_same_pci_cells = self.target_cells[
            (self.target_cells['pci'] == candidate_pci) &
            (self.target_cells['pci'].notna()) &
            (self.target_cells['earfcn_dl'] == target_earfcn) &  # 同频条件
            (self.target_cells['earfcn_dl'].notna()) &
            (self.target_cells['cell_type'] == self.network_type)  # 同网络类型
        ].copy()
        
        # 排除当前小区自己
        if exclude_enodeb is not None and exclude_cell is not None:
            same_freq_same_pci_cells = same_freq_same_pci_cells[
                ~((same_freq_same_pci_cells['enodeb_id'] == exclude_enodeb) & 
                  (same_freq_same_pci_cells['cell_id'] == exclude_cell))
            ]
        
        # 过滤有效位置的小区
        same_freq_same_pci_cells = same_freq_same_pci_cells.dropna(subset=['lat', 'lon'])
        
        if same_freq_same_pci_cells.empty:
            result = (True, float('inf'))
            self.pci_validity_cache[cache_key] = result
            return result
        
        # 计算到所有同频同PCI小区的距离
        distances = self.calculate_distance_vectorized(
            target_lat, target_lon,
            same_freq_same_pci_cells['lat'].values,
            same_freq_same_pci_cells['lon'].values
        )
        
        min_distance = np.min(distances)
        is_valid = min_distance >= self.reuse_distance_km
        
        result = (is_valid, min_distance)
        self.pci_validity_cache[cache_key] = result
        return result
    
    def get_same_site_cells(self, target_lat: float, target_lon: float, exclude_enodeb_id: int = None, exclude_cell_id: int = None) -> List[Dict]:
        """
        获取同站点的其他小区信息
        真正基于经纬度判断同站点，不依赖eNodeBID
        """
        if pd.isna(target_lat) or pd.isna(target_lon):
            return []
        
        # 使用经纬度查找相同位置的小区
        # 容差值设为0.0001度（约10米精度），确保同一物理位置的小区能够正确匹配
        tolerance = 0.0001  # 约10米精度
        location_filter = (
            (abs(self.all_cells_combined['lat'] - target_lat) < tolerance) &
            (abs(self.all_cells_combined['lon'] - target_lon) < tolerance)
        )
        
        # 从所有小区数据中查找相同位置的小区（包括未分配PCI的小区）
        same_site_cells = self.all_cells_combined[location_filter].copy()
        
        # 排除指定的基站ID和小区ID
        if exclude_enodeb_id is not None and exclude_cell_id is not None:
            same_site_cells = same_site_cells[
                ~((same_site_cells['enodeb_id'] == exclude_enodeb_id) & 
                  (same_site_cells['cell_id'] == exclude_cell_id))
            ]
        
        # 只返回当前网络类型的小区
        same_site_cells = same_site_cells[
            same_site_cells['cell_type'] == self.network_type
        ]
        
        # 确保获取到最新的PCI分配信息
        result = same_site_cells.to_dict('records')
        
        # 调试信息：显示同站点小区的PCI分配情况
        if result:
            print(f"      [DEBUG] 位置({target_lat:.6f}, {target_lon:.6f})同站点小区PCI分配情况:")
            for cell in result:
                enodeb_id = cell.get('enodeb_id', 'N/A')
                cell_id = cell.get('cell_id', 'N/A')
                pci = cell.get('pci', 'N/A')
                mod_val = int(pci) % self.mod_value if pd.notna(pci) and pci != -1 else 'N/A'
                print(f"        基站{enodeb_id}-小区{cell_id}: PCI={pci}, mod{self.mod_value}={mod_val}")
        
        return result
    
    def get_cells_at_same_location(self, target_lat: float, target_lon: float, 
                                  exclude_enodeb_id: int = None, exclude_cell_id: int = None) -> List[Dict]:
        """
        获取相同经纬度位置的其他小区信息
        
        Args:
            target_lat: 目标纬度
            target_lon: 目标经度
            exclude_enodeb_id: 要排除的基站ID
            exclude_cell_id: 要排除的小区ID
            
        Returns:
            相同位置的小区列表
        """
        # 使用经纬度查找相同位置的小区
        # 增大容差值到0.0001度（约10米精度），确保同一基站的小区能够正确匹配
        tolerance = 0.0001  # 约10米精度
        location_filter = (
            (abs(self.all_cells_combined['lat'] - target_lat) < tolerance) &
            (abs(self.all_cells_combined['lon'] - target_lon) < tolerance)
        )
        
        # 从所有小区数据中查找相同位置的小区（包括未分配PCI的小区）
        # 关键修改：统一使用all_cells_combined作为数据源，确保同站点判断一致性
        same_location_cells = self.all_cells_combined[location_filter]
        
        # 排除指定的基站ID
        if exclude_enodeb_id is not None:
            same_location_cells = same_location_cells[
                same_location_cells['enodeb_id'] != exclude_enodeb_id
            ]
        
        # 排除指定的小区ID
        if exclude_cell_id is not None:
            same_location_cells = same_location_cells[
                same_location_cells['cell_id'] != exclude_cell_id
            ]
        
        # 关键修改：只返回当前网络类型的小区，确保与get_same_site_cells函数逻辑一致
        same_location_cells = same_location_cells[
            same_location_cells['cell_type'] == self.network_type
        ]
        
        return same_location_cells.to_dict('records')
    
    def check_same_site_mod_conflict(self, candidate_pci: int, target_lat: float, target_lon: float,
                                   exclude_enodeb_id: int = None, exclude_cell_id: int = None) -> bool:
        """
        检查候选PCI是否与同站点小区的模值冲突
        真正基于经纬度判断同站点
        支持NR网络的双模约束（同时检查模3和模30冲突）

        Args:
            candidate_pci: 候选PCI值
            target_lat: 目标小区纬度
            target_lon: 目标小区经度
            exclude_enodeb_id: 要排除的基站ID
            exclude_cell_id: 要排除的小区ID

        Returns:
            True表示无冲突，False表示有冲突
        """
        candidate_mod = candidate_pci % self.mod_value

        # 检查同站点小区模值冲突（基于经纬度判断）
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb_id, exclude_cell_id)

        # 统计同站点已有的模值
        existing_mods = []
        conflict_cells = []

        # 检查是否与同站点其他小区的模值冲突
        for cell in same_site_cells:
            existing_pci = cell['pci']
            if pd.notna(existing_pci) and existing_pci != -1:
                existing_mod = int(existing_pci) % self.mod_value
                existing_mods.append(existing_mod)
                if existing_mod == candidate_mod:
                    enodeb_id = cell.get('enodeb_id', 'N/A')
                    cell_id = cell.get('cell_id', 'N/A')
                    conflict_cells.append(f"基站{enodeb_id}-小区{cell_id}(PCI{existing_pci})")

        # 关键修复：无论是否继承模值，都必须避免同站同模冲突
        # 对于LTE网络，同站3个小区必须保证模3值不同
        if self.network_type == "LTE":
            # 检查同站点是否已有3个不同模值
            unique_mods = set(existing_mods)
            if len(unique_mods) >= 3 and candidate_mod in unique_mods:
                # 同站点已经有3个不同模值，且候选PCI的模值已经存在，禁止分配
                print(f"      [DEBUG] 发现LTE同站点模{self.mod_value}冲突: 候选PCI={candidate_pci} (mod{self.mod_value}={candidate_mod})")
                print(f"      [DEBUG] 同站点已有模{self.mod_value}值: {sorted(existing_mods)} (已达3种不同值)")
                return False

        # 对于NR网络，需要同时检查模3和模30冲突
        if self.network_type == "NR" and self.dual_mod_requirement:
            # 检查模3冲突
            candidate_mod3 = candidate_pci % 3
            existing_mod3s = []
            mod3_conflict_cells = []
            
            for cell in same_site_cells:
                existing_pci = cell['pci']
                if pd.notna(existing_pci) and existing_pci != -1:
                    existing_mod3 = int(existing_pci) % 3
                    existing_mod3s.append(existing_mod3)
                    if existing_mod3 == candidate_mod3:
                        enodeb_id = cell.get('enodeb_id', 'N/A')
                        cell_id = cell.get('cell_id', 'N/A')
                        mod3_conflict_cells.append(f"基站{enodeb_id}-小区{cell_id}(PCI{existing_pci})")
            
            # 检查同站点是否已有3个不同模3值
            unique_mod3s = set(existing_mod3s)
            if len(unique_mod3s) >= 3 and candidate_mod3 in unique_mod3s:
                print(f"      [DEBUG] 发现NR同站点模3冲突: 候选PCI={candidate_pci} (mod3={candidate_mod3})")
                print(f"      [DEBUG] 同站点已有模3值: {sorted(existing_mod3s)} (已达3种不同值)")
                return False
            
            # 检查模3冲突
            if mod3_conflict_cells:
                print(f"      [DEBUG] 发现NR同站点模3冲突: 候选PCI={candidate_pci} (mod3={candidate_mod3})")
                print(f"      [DEBUG] 模3冲突小区: {', '.join(mod3_conflict_cells)}")
                print(f"      [DEBUG] 同站点已有模3值: {sorted(existing_mod3s)}")
                return False

        # 如果发现冲突，记录详细信息
        if conflict_cells:
            print(f"      [DEBUG] 发现同站点模{self.mod_value}冲突: 候选PCI={candidate_pci} (mod{self.mod_value}={candidate_mod})")
            print(f"      [DEBUG] 冲突小区: {', '.join(conflict_cells)}")
            print(f"      [DEBUG] 同站点已有模{self.mod_value}值: {sorted(existing_mods)}")
            return False  # 发现冲突

        return True  # 无冲突

    def get_reuse_compliant_pcis(self, target_lat: float, target_lon: float, target_earfcn: float,
                                exclude_enodeb: int = None, exclude_cell: int = None,
                                target_mod: Optional[int] = None) -> List[Tuple[int, float]]:
        """
        获取满足复用距离要求的PCI列表，按照多重优先级排序
        增强同站点模值冲突避免逻辑
        
        Returns:
            List of (pci, min_distance) tuples, sorted by preference
        """
        compliant_pcis = []
        
        # 如果需要继承mod值，严格按照mod值筛选PCI范围
        if self.inherit_mod and target_mod is not None:
            # 严格继承模式：只考虑匹配mod值的PCI
            candidate_pcis = [pci for pci in self.pci_range if pci % self.mod_value == target_mod]
            print(f"    严格mod{self.mod_value}继承模式: 只检查mod{self.mod_value}={target_mod}的PCI，候选PCI数量: {len(candidate_pcis)}")
        else:
            # 自由规划模式：考虑所有PCI，不考虑mod值
            candidate_pcis = self.pci_range
            print(f"    自由规划模式: 检查所有PCI，不考虑mod{self.mod_value}值，候选PCI数量: {len(candidate_pcis)}")
        
        # 获取同站点小区的模值信息，用于避免冲突
        same_site_mods = set()
        # 关键修复：直接基于经纬度获取同站点小区，不依赖eNodeBID
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
        for cell in same_site_cells:
            # 增强：考虑同站点的所有小区，包括不同基站但同位置的小区
            # 优先级：已分配PCI > 原PCI > 忽略无PCI的小区
            pci_value = cell.get('pci')
            original_pci = cell.get('原PCI')

            # 首先检查已分配的PCI
            if pd.notna(pci_value) and pci_value != -1:
                same_site_mods.add(int(pci_value) % self.mod_value)
            # 如果没有分配PCI，检查原PCI
            elif pd.notna(original_pci) and original_pci != -1:
                same_site_mods.add(int(original_pci) % self.mod_value)
            # 如果都没有PCI值，则跳过该小区

        print(f"    同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")

        # 关键修复：对于LTE网络，即使自由规划模式也要考虑同站点模3冲突避免
        if self.network_type == "LTE" and not self.inherit_mod:
            print(f"    LTE自由规划模式：将严格避免同站点模3冲突")
            # 检查同站点是否已经有3个不同模值
            if len(same_site_mods) >= 3:
                print(f"    警告：同站点已有3种不同模{self.mod_value}值，将强制选择不同模值")
        
        # 对于NR网络，需要同时考虑模3和模30约束
        if self.network_type == "NR" and self.dual_mod_requirement:
            print(f"    NR双模约束模式：将同时避免同站点模3和模30冲突")
            # 统计同站点已有的模3值
            same_site_mod3s = set()
            for cell in same_site_cells:
                pci_value = cell.get('pci')
                original_pci = cell.get('原PCI')
                
                # 首先检查已分配的PCI
                if pd.notna(pci_value) and pci_value != -1:
                    same_site_mod3s.add(int(pci_value) % 3)
                # 如果没有分配PCI，检查原PCI
                elif pd.notna(original_pci) and original_pci != -1:
                    same_site_mod3s.add(int(original_pci) % 3)
            
            print(f"    同站点已有模3值: {sorted(list(same_site_mod3s))}")
            
            # 检查同站点是否已经有3个不同模3值
            if len(same_site_mod3s) >= 3:
                print(f"    警告：同站点已有3种不同模3值，将强制选择不同模3值")
        
        # 检查候选PCI
        for pci in candidate_pcis:
            # 1. 首先检查复用距离（最高优先级）
            is_valid, min_distance = self.validate_pci_reuse_distance(
                pci, target_lat, target_lon, target_earfcn, exclude_enodeb, exclude_cell
            )
            
            if not is_valid:
                continue  # 不满足复用距离要求的PCI直接跳过
            
            # 2. 检查同站点模值冲突（第二优先级）
            candidate_mod = pci % self.mod_value
            has_same_site_mod_conflict = candidate_mod in same_site_mods
            
            # 3. 计算PCI分布均衡性指标（第三优先级）
            # 距离越接近阈值越均衡
            if min_distance == float('inf'):
                balance_score = 0  # 无复用PCI，最均衡
            else:
                # 距离接近阈值的得分更高（分数越小越好）
                balance_score = abs(min_distance - self.reuse_distance_km)
            
            compliant_pcis.append((pci, min_distance, has_same_site_mod_conflict, balance_score))
        
        if not compliant_pcis:
            return []

        # 关键修复：对于LTE网络，无论是否继承模值，都必须严格避免同站同模
        if self.network_type == "LTE":
            # LTE网络强制同站点模3错开原则
            print(f"    LTE网络：强制执行同站点模3错开原则")

            # 按模3值分组
            mod_groups = {}
            for pci_info in compliant_pcis:
                pci, distance, has_mod_conflict, balance_score = pci_info
                mod_value = pci % self.mod_value
                if mod_value not in mod_groups:
                    mod_groups[mod_value] = []
                mod_groups[mod_value].append(pci_info)

            # 获取同站点已分配的模3值
            same_site_assigned_mods = set()
            same_site_cells_for_check = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
            for cell in same_site_cells_for_check:
                if 'pci' in cell and pd.notna(cell['pci']) and cell['pci'] != -1:
                    same_site_assigned_mods.add(int(cell['pci']) % self.mod_value)

            print(f"    同站点已有模3值: {sorted(list(same_site_assigned_mods))}")
            print(f"    候选PCI模3值分布: {sorted(list(mod_groups.keys()))}")

            # 关键修复：LTE网络必须保证同站3个小区模3值不同
            # 如果同站点已有2个不同模3值，必须选择第3个不同的模3值
            if len(same_site_assigned_mods) >= 2:
                print(f"    同站点已有{len(same_site_assigned_mods)}个不同模3值，必须选择不同的模3值")
                # 强制选择未被使用的模3值
                available_mod_groups = {mod: pci_list for mod, pci_list in mod_groups.items()
                                      if mod not in same_site_assigned_mods}
                
                if available_mod_groups:
                    # 有可用模3值，优先选择
                    final_pcis = []
                    for mod_value, pci_list in available_mod_groups.items():
                        # 对每个模3组内的PCI进行排序（复用距离优先）
                        pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                        final_pcis.extend(pci_list)
                    print(f"    成功找到{len(final_pcis)}个无同站点模3冲突的PCI")
                else:
                    # 所有模3值都被使用，这是严重问题 - 需要强制选择不同的模3值
                    print(f"    严重警告：所有模3值都已被同站点小区使用，违反LTE同站3小区模3错开原则！")
                    
                    # 统计每个mod值的使用次数
                    mod_usage_count = {}
                    for mod in same_site_assigned_mods:
                        mod_usage_count[mod] = 0

                    # 计算每个mod值在同站点的使用次数
                    for cell in same_site_cells_for_check:
                        if 'pci' in cell and pd.notna(cell['pci']) and cell['pci'] != -1:
                            cell_mod = int(cell['pci']) % self.mod_value
                            if cell_mod in mod_usage_count:
                                mod_usage_count[cell_mod] += 1

                    # 找到使用次数最少的mod值
                    if mod_usage_count:
                        min_usage_mod = min(mod_usage_count.keys(), key=lambda x: mod_usage_count[x])
                        print(f"    使用次数最少的mod值: {min_usage_mod} (使用次数: {mod_usage_count[min_usage_mod]})")

                        # 优先选择使用次数最少的mod值的PCI
                        if min_usage_mod in mod_groups:
                            best_mod_pci_list = mod_groups[min_usage_mod]
                            best_mod_pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                            final_pcis = best_mod_pci_list
                        else:
                            # 如果没有找到，从所有mod组中选择
                            final_pcis = []
                            for mod_value, pci_list in mod_groups.items():
                                pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                                final_pcis.extend(pci_list)
                    else:
                        # 如果没有使用统计，从所有mod组中选择
                        final_pcis = []
                        for mod_value, pci_list in mod_groups.items():
                            pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                            final_pcis.extend(pci_list)
            else:
                # 同站点模3值不足2个，可以自由选择但优先避免冲突
                available_mod_groups = {mod: pci_list for mod, pci_list in mod_groups.items()
                                      if mod not in same_site_assigned_mods}
                
                if available_mod_groups:
                    # 优先选择未被使用的模3值
                    final_pcis = []
                    for mod_value, pci_list in available_mod_groups.items():
                        pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                        final_pcis.extend(pci_list)
                    print(f"    成功找到{len(final_pcis)}个无同站点模3冲突的PCI")
                else:
                    # 如果没有可用模3值，使用所有候选PCI
                    final_pcis = []
                    for mod_value, pci_list in mod_groups.items():
                        pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                        final_pcis.extend(pci_list)
                    print(f"    使用所有候选PCI，共{len(final_pcis)}个")
        else:
            # NR网络或非LTE网络的原有逻辑
            # 新优先级策略：复用距离 > 同站点模值冲突避免 > PCI分布均衡性
            # 首先尝试无同站点模值冲突的PCI（最高优先级）
            no_mod_conflict_pcis = [(pci, distance, has_mod_conflict, balance_score)
                                   for pci, distance, has_mod_conflict, balance_score in compliant_pcis
                                   if not has_mod_conflict]

            # 对于NR网络的双模约束，需要额外检查模3冲突
            if self.network_type == "NR" and self.dual_mod_requirement:
                print(f"    NR双模约束：将同时检查模30和模3冲突")
                
                # 统计同站点已有的模3值
                same_site_mod3s = set()
                same_site_cells_for_check = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
                for cell in same_site_cells_for_check:
                    if 'pci' in cell and pd.notna(cell['pci']) and cell['pci'] != -1:
                        same_site_mod3s.add(int(cell['pci']) % 3)
                
                print(f"    同站点已有模3值: {sorted(list(same_site_mod3s))}")
                
                # 过滤掉有模3冲突的PCI
                no_mod3_conflict_pcis = []
                for pci_info in no_mod_conflict_pcis:
                    pci = pci_info[0]
                    mod3_value = pci % 3
                    if mod3_value not in same_site_mod3s:
                        no_mod3_conflict_pcis.append(pci_info)
                
                if no_mod3_conflict_pcis:
                    print(f"    成功找到{len(no_mod3_conflict_pcis)}个无模30和模3冲突的PCI")
                    final_pcis = no_mod3_conflict_pcis
                else:
                    print(f"    警告：所有无模30冲突的PCI都存在模3冲突")
                    # 如果没有无模3冲突的PCI，使用所有无模30冲突的PCI
                    final_pcis = no_mod_conflict_pcis
                    print(f"    使用{len(final_pcis)}个无模30冲突但可能存在模3冲突的PCI")
            else:
                 # 非NR双模约束或非NR网络，使用原有逻辑
                 if no_mod_conflict_pcis:
                     # 优先使用无模值冲突的PCI
                     final_pcis = no_mod_conflict_pcis
                     print(f"    成功找到{len(final_pcis)}个无同站点模{self.mod_value}冲突的PCI")
                 else:
                     # 关键修复：即使所有PCI都有模3冲突，也要优先选择模3值不同的PCI
                     # 避免同站所有小区模3值相同的情况
                     print(f"    警告：所有满足复用距离的PCI都与同站点存在模{self.mod_value}冲突")
                     print(f"    同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")

                 # 按模3值分组，优先选择模3值不同的PCI
                 mod_groups = {}
                 for pci_info in compliant_pcis:
                     pci, distance, has_mod_conflict, balance_score = pci_info
                     mod_value = pci % self.mod_value
                     if mod_value not in mod_groups:
                         mod_groups[mod_value] = []
                     mod_groups[mod_value].append(pci_info)

                 # 关键修复：绝对不允许同站同模，必须选择不同的模3值
                 if len(mod_groups) > 1:
                     print(f"    发现{len(mod_groups)}个不同的模{self.mod_value}值组，将强制选择不同的模3值")

                     # 获取当前位置已分配PCI的模3值，避免重复选择
                     same_site_assigned_mods = set()
                     # 基于经纬度获取同站点已分配的小区PCI模3值
                     same_site_cells_for_check = self.get_same_site_cells(target_lat, target_lon, exclude_enodeb, exclude_cell)
                     for cell in same_site_cells_for_check:
                         if 'pci' in cell and pd.notna(cell['pci']):
                             same_site_assigned_mods.add(int(cell['pci']) % self.mod_value)

                     # 强制选择未被同基站其他小区使用的模3值
                     available_mod_groups = {mod: pci_list for mod, pci_list in mod_groups.items()
                                           if mod not in same_site_assigned_mods}

                     if available_mod_groups:
                         # 优先选择未被使用的模3值
                         final_pcis = []
                         for mod_value, pci_list in available_mod_groups.items():
                             # 对每个模3组内的PCI进行排序（复用距离优先）
                             pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                             final_pcis.append(pci_list[0])
                     else:
                         # 如果所有模3值都被使用，必须选择不同的模3值来避免同站同模
                         # 选择与已分配模3值不同的模3值
                         remaining_mod_groups = {mod: pci_list for mod, pci_list in mod_groups.items()}
                         if len(remaining_mod_groups) > 1:
                             # 选择与已分配模3值不同的模3值
                             final_pcis = []
                             for mod_value, pci_list in remaining_mod_groups.items():
                                 if mod_value not in same_site_assigned_mods:
                                     pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                                     final_pcis.append(pci_list[0])
                             if not final_pcis:
                                 # 如果无法避免，选择最优的PCI，但记录严重警告
                                 print(f"    严重警告：无法避免同站同模！将选择最优PCI")
                                 final_pcis = []
                                 for mod_value, pci_list in mod_groups.items():
                                     pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                                     final_pcis.append(pci_list[0])
                         else:
                             # 如果只有一个模3值，无法避免同站同模
                             print(f"    严重警告：所有候选PCI模{self.mod_value}值相同，无法避免同站同模！")
                             # 强制选择不同的PCI，即使模3值相同也要避免同站同模
                             final_pcis = []
                             for mod_value, pci_list in mod_groups.items():
                                 pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                                 final_pcis.append(pci_list[0])
                 else:
                     # 如果所有PCI模3值都相同，无法避免同站同模
                     print(f"    严重警告：所有候选PCI模{self.mod_value}值相同，无法避免同站同模！")
                     # 强制选择不同的PCI，即使模3值相同也要避免同站同模
                     final_pcis = []
                     for mod_value, pci_list in mod_groups.items():
                         pci_list.sort(key=lambda x: (-x[1], x[3], x[0]))
                         final_pcis.append(pci_list[0])
        
        # 多重优先级排序策略 - 修正优先级顺序
        def sort_key(pci_info):
            pci, distance, has_mod_conflict, balance_score = pci_info
            
            # 优先级1：同站点模值冲突（最高优先级）- 绝对不允许同站同模
            mod_conflict_priority = 0 if not has_mod_conflict else 1
            
            # 优先级2：复用距离合规性（第二优先级）- 必须满足用户设定的最小复用距离
            # 关键修复：优先选择满足复用距离要求的PCI
            if distance == float('inf'):
                distance_compliance = 0  # 无复用PCI最优
            elif distance >= self.reuse_distance_km:
                distance_compliance = 0  # 满足复用距离要求
            else:
                distance_compliance = 1  # 不满足复用距离要求
            
            # 优先级3：复用距离大小（第三优先级）- 距离越大越安全
            if distance == float('inf'):
                distance_priority = -999999  # 无复用PCI最优
            else:
                distance_priority = -distance  # 距离大的优先
            
            # 优先级4：PCI分布均衡性（第四优先级）- 越接近阈值越好
            balance_priority = balance_score
            
            # 优先级5：PCI值（第五优先级）- 小的优先，确保确定性
            pci_priority = pci
            
            return (mod_conflict_priority, distance_compliance, distance_priority, balance_priority, pci_priority)
        
        final_pcis.sort(key=sort_key)
        
        # 输出排序结果用于调试
        if len(final_pcis) > 0:
            top_candidates = final_pcis[:min(5, len(final_pcis))]
            print(f"    前{len(top_candidates)}个候选PCI排序结果:")
            for i, (pci, dist, conflict, _) in enumerate(top_candidates):
                status = "无冲突" if not conflict else "有冲突"
                print(f"      {i+1}. PCI={pci} (mod{self.mod_value}={pci%self.mod_value}), 距离={dist:.2f}km, {status}")
        
        # 保持4元组格式以确保数据结构一致性
        return final_pcis
    
    def assign_pci_with_reuse_priority(self, enodeb_id: int, cell_id: int) -> Tuple[Optional[int], str, float, float]:
        """
        优先确保复用距离的PCI分配方法

        Returns:
            (assigned_pci, reason, earfcn_dl, actual_min_reuse_distance)
            assigned_pci: 分配的PCI值，如果无法分配则为None
        """
        # 获取小区信息
        cell_info, status = self.get_cell_info(enodeb_id, cell_id)

        if status != 'success':
            # 对于找不到的小区或缺少位置信息的小区，不进行PCI分配
            if status == 'cell_not_found':
                # 小区未找到，返回None表示不分配PCI
                return None, 'cell_not_found_no_pci', 0.0, 0.0
            elif status == 'no_location':
                # 缺少位置信息，不进行PCI分配，按优化要求返回特定格式
                earfcn_dl = cell_info['earfcn_dl'] if cell_info is not None else 0.0
                return None, 'no_location_fallback', earfcn_dl, 0.0
        
        target_lat = cell_info['lat']
        target_lon = cell_info['lon']
        earfcn_dl = cell_info['earfcn_dl']
        
        # 确定目标mod值（如果需要继承）
        target_mod = None
        if self.inherit_mod and not pd.isna(cell_info['pci']):
            target_mod = int(cell_info['pci']) % self.mod_value
        
        # 获取满足复用距离的PCI列表（同频检查）- 现在直接返回4元组格式
        compliant_pcis = self.get_reuse_compliant_pcis(
            target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
        )
        # 关键修复：对于LTE网络，不再进行额外的模3均衡化处理，避免破坏同站点模3检查逻辑
        # 因为get_reuse_compliant_pcis方法已经包含了完善的LTE同站点模3错开逻辑
        if self.network_type != "LTE" and len(compliant_pcis) > 0:
            # 仅对非LTE网络进行模3均衡化处理
            # 按模3值分组
            mod_groups = {}
            for pci_info in compliant_pcis:
                pci = pci_info[0]
                mod = pci % self.mod_value
                if mod not in mod_groups:
                    mod_groups[mod] = []
                mod_groups[mod].append(pci_info)
            
            # 确保每个模3组最多保留3个最佳候选
            balanced_pcis = []
            for mod in mod_groups:
                # 按距离和均衡性排序（4元组结构）
                mod_groups[mod].sort(key=lambda x: (-x[1], x[3]))
                balanced_pcis.extend(mod_groups[mod][:3])
            
            # 如果不同模3值不足3组，放宽距离要求
            if len(mod_groups) < 3 and len(compliant_pcis) > 20:
                print(f"  警告：候选PCI模3多样性不足({len(mod_groups)}种)，将放宽距离要求")
                # 临时放宽距离要求
                original_distance = self.reuse_distance_km
                self.reuse_distance_km = 2.0  # 临时放宽距离要求
                compliant_pcis = self.get_reuse_compliant_pcis(
                    target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
                )
                self.reuse_distance_km = original_distance  # 恢复原距离要求
                # 重新平衡
                mod_groups = {}
                for pci_info in compliant_pcis:
                    pci = pci_info[0]
                    mod = pci % self.mod_value
                    if mod not in mod_groups:
                        mod_groups[mod] = []
                    mod_groups[mod].append(pci_info)
                balanced_pcis = []
                for mod in mod_groups:
                    mod_groups[mod].sort(key=lambda x: (-x[1], x[3] if len(x) > 3 else 0))
                    balanced_pcis.extend(mod_groups[mod][:3])
            
            compliant_pcis = balanced_pcis
            print(f"  模3均衡化处理后候选PCI数量: {len(compliant_pcis)} (含{len(mod_groups)}种模3值)")
        else:
            # LTE网络直接使用get_reuse_compliant_pcis返回的结果，不进行额外处理
            print(f"  LTE网络：使用get_reuse_compliant_pcis的完整结果，共{len(compliant_pcis)}个候选PCI")
            
            # 对于LTE网络，如果候选PCI数量较少但大于0，直接使用所有候选
            if len(compliant_pcis) > 0 and len(compliant_pcis) <= 10:
                print(f"  候选PCI数量较少({len(compliant_pcis)})，直接使用所有候选确保同站模3错开")
        
        if not compliant_pcis:
            # 智能降级策略：当无法找到满足复用距离的PCI时，逐步放宽要求
            print(f"  警告：无法找到满足{self.reuse_distance_km}km复用距离的PCI，尝试智能降级")
            
            # 第一步：尝试放宽到3.0km（中等距离要求）
            original_distance = self.reuse_distance_km
            self.reuse_distance_km = 3.0
            compliant_pcis = self.get_reuse_compliant_pcis(
                target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
            )
            
            if not compliant_pcis:
                # 第二步：尝试放宽到2.0km（较低距离要求）
                self.reuse_distance_km = 2.0
                compliant_pcis = self.get_reuse_compliant_pcis(
                    target_lat, target_lon, earfcn_dl, enodeb_id, cell_id, target_mod
                )
            
            # 恢复原距离要求
            self.reuse_distance_km = original_distance
            
            if compliant_pcis:
                print(f"  智能降级成功：找到{len(compliant_pcis)}个候选PCI")
                # 使用降级后的候选PCI继续规划
                best_pci_info = compliant_pcis[0]
                best_pci = best_pci_info[0]
                min_distance = best_pci_info[1]
                
                # 确定分配原因
                reason = f'no_compliant_pci_fallback_downgrade_{self.reuse_distance_km}km_to_{min(3.0, 2.0)}km'
                
                return best_pci, reason, earfcn_dl, min_distance
            else:
                # 如果降级后仍然找不到，使用保底方案
                self.failure_reasons['no_compliant_pci'].append(f"{enodeb_id}-{cell_id}")
                fallback_pci = enodeb_id % len(self.pci_range)
                return fallback_pci, 'no_compliant_pci_fallback', earfcn_dl, 0.0
        
        # 选择最优的PCI（经过多重优先级排序后的第一个）
        best_pci_info = compliant_pcis[0]
        best_pci = best_pci_info[0]
        min_distance = best_pci_info[1]
        has_conflict = best_pci_info[2] if len(best_pci_info) > 2 else False
        balance_score = best_pci_info[3] if len(best_pci_info) > 3 else 0

        # 检查选择的PCI的特性
        best_pci_mod = best_pci % self.mod_value

        # 获取同站点小区的模值信息用于详细显示
        same_site_mods = set()
        same_site_cells = self.get_same_site_cells(target_lat, target_lon, enodeb_id, cell_id)
        for cell in same_site_cells:
            # 只考虑已经分配了PCI的小区（pci不为空且不为-1）
            if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                same_site_mods.add(int(cell['pci']) % self.mod_value)

        same_site_mod_ok = best_pci_mod not in same_site_mods
        
        # 显示PCI选择的详细信息
        print(f"    小区{enodeb_id}-{cell_id}: 候选PCI数量={len(compliant_pcis)}")
        print(f"    同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")
        
        if len(compliant_pcis) > 1:
            top_candidates = compliant_pcis[:min(3, len(compliant_pcis))]
            print(f"    前{len(top_candidates)}个候选PCI: {[(pci, f'{dist:.2f}km') for pci, dist, _, _ in top_candidates]}")
        
        print(f"    选择PCI={best_pci} (mod{self.mod_value}={best_pci_mod})")
        print(f"      复用距离: {min_distance:.2f}km")
        print(f"      同站点模{self.mod_value}冲突: {'无' if same_site_mod_ok else '有'}")
        
        if not same_site_mod_ok:
            print(f"      警告: 与同站点小区模{self.mod_value}值冲突!")
            conflict_cells = []
            for cell in same_site_cells:
                # 只考虑已经分配了PCI的小区（pci不为空且不为-1）
                if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                    cell_mod = int(cell['pci']) % self.mod_value
                    if cell_mod == best_pci_mod:
                        conflict_cells.append(f"小区{cell['cell_id']}(PCI{cell['pci']})")
            if conflict_cells:
                print(f"      冲突小区: {', '.join(conflict_cells)}")
        
        if min_distance != float('inf'):
            if abs(min_distance - self.reuse_distance_km) <= 2:
                print(f"      均衡性: 优秀 (接近{self.reuse_distance_km}km阈值)")
            elif min_distance <= self.reuse_distance_km + 5:
                print(f"      均衡性: 良好 (适中距离)")
            else:
                print(f"      均衡性: 一般 (距离较远)")
        
        # 确定分配原因
        if self.inherit_mod and target_mod is not None:
            # 继承模式：由于严格筛选，分配的PCI必然匹配mod值
            if self.network_type == "LTE":
                reason = 'strict_mod3_inheritance'
            else:
                reason = 'strict_mod30_inheritance'
        else:
            # 自由规划模式：仅考虑复用距离
            reason = 'free_planning_reuse_compliant'
        
        return best_pci, reason, earfcn_dl, min_distance
    
    def calculate_pci_mod(self, pci_value) -> Optional[int]:
        """
        计算PCI的模值（LTE为mod3，NR为mod30）
        """
        if pci_value is None or pci_value == -1 or pd.isna(pci_value):
            return None
        try:
            return int(pci_value) % self.mod_value
        except (ValueError, TypeError):
            return None
    
    def update_cell_pci(self, enodeb_id: int, cell_id: int, assigned_pci: Optional[int]):
        """
        更新小区的PCI值
        
        Args:
            assigned_pci: 分配的PCI值，如果为None表示无法分配PCI
        """
        if assigned_pci is None:
            # 无法分配PCI，不进行更新
            print(f"      [DEBUG] 无法为基站{enodeb_id}-小区{cell_id}分配PCI，跳过更新")
            return
        
        mask = (self.target_cells['enodeb_id'] == enodeb_id) & (self.target_cells['cell_id'] == cell_id)
        self.target_cells.loc[mask, 'pci'] = assigned_pci
        
        # 同时更新合并数据
        if self.all_cells_combined is not None:
            combined_mask = ((self.all_cells_combined['enodeb_id'] == enodeb_id) & 
                           (self.all_cells_combined['cell_id'] == cell_id) & 
                           (self.all_cells_combined['cell_type'] == self.network_type))
            self.all_cells_combined.loc[combined_mask, 'pci'] = assigned_pci
        
        # 彻底清除所有相关缓存（关键修复）
        # 1. 清除同站点缓存
        same_site_keys = [key for key in self.same_site_cache.keys() if key[0] == enodeb_id]
        for key in same_site_keys:
            del self.same_site_cache[key]
        
        # 2. 清除距离缓存（避免使用旧的干扰计算结果）
        cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
        distance_keys = []  # 初始化变量，提供安全默认值
        if cell_info is not None and not pd.isna(cell_info['lat']) and not pd.isna(cell_info['lon']):
            distance_keys = [k for k in self.distance_cache.keys()
                           if k[0] == round(cell_info['lat'], 6)
                           and k[1] == round(cell_info['lon'], 6)]
            for key in distance_keys:
                del self.distance_cache[key]
            
        # 3. 清除PCI验证缓存
        pci_keys = [k for k in self.pci_validity_cache.keys()
                   if k[3] == enodeb_id or k[4] == cell_id]
        for key in pci_keys:
            del self.pci_validity_cache[key]
            
        print(f"      [DEBUG] 更新基站{enodeb_id}-小区{cell_id}的PCI为{assigned_pci}，清除缓存:")
        print(f"        - 同站点缓存: {len(same_site_keys)}项")
        print(f"        - 距离缓存: {len(distance_keys)}项")
        print(f"        - PCI验证缓存: {len(pci_keys)}项")
    
    def calculate_final_min_reuse_distance(self, result_df: pd.DataFrame) -> List:
        """
        计算每个小区的最终最小PCI复用距离（验证分配结果）
        返回混合类型列表，数值距离保持为float，特殊情况使用字符串标记
        """
        print("\\n正在验证最终复用距离...")
        min_reuse_distances = []

        for idx, cell in result_df.iterrows():
            # 根据网络类型获取正确的列名
            if self.network_type == "NR":
                enodeb_id = cell.get('gNodeBID')
            else:
                enodeb_id = cell.get('eNodeBID')
            cell_id = cell.get('CellID')
            assigned_pci = cell.get('分配的PCI')

            # 处理PCI分配为None的情况（无位置信息的小区）
            if assigned_pci is None:
                min_reuse_distances.append("位置信息缺失")
                continue

            if assigned_pci == -1:
                min_reuse_distances.append("分配失败")
                continue

            # 获取当前小区信息
            cell_info, status = self.get_cell_info(enodeb_id, cell_id)

            if status != 'success' or pd.isna(cell_info['lat']) or pd.isna(cell_info['lon']):
                min_reuse_distances.append("位置信息缺失")
                continue

            # 验证实际复用距离（同频检查）
            _, actual_min_distance = self.validate_pci_reuse_distance(
                assigned_pci, cell_info['lat'], cell_info['lon'], cell_info['earfcn_dl'], enodeb_id, cell_id
            )

            # 保持数值格式，特殊情况使用字符串
            if actual_min_distance == float('inf'):
                min_reuse_distances.append("无复用PCI")
            else:
                min_reuse_distances.append(round(actual_min_distance, 2))

            if (idx + 1) % 100 == 0 or (idx + 1) == len(result_df):
                print(f"已验证 {idx+1}/{len(result_df)} 个小区的复用距离")

        return min_reuse_distances
    
    def plan_pci_with_reuse_priority(self) -> pd.DataFrame:
        """
        执行优先确保复用距离的PCI规划
        """
        if self.cells_to_plan is None:
            raise ValueError("请先加载数据")
        
        start_time = time.time()
        
        print(f"\\n开始{self.network_type}网络PCI规划（优先确保同频PCI复用距离）")
        print(f"同频PCI最小复用距离: {self.reuse_distance_km}km (优先级最高)")
        print(f"PCI范围: 0-{max(self.pci_range)}")
        
        if self.network_type == "LTE":
            if self.lte_inherit_mod3:
                print(f"规划模式: 严格mod3继承 - 只分配匹配原PCI mod3值的PCI")
            else:
                print(f"规划模式: 自由规划 - 不考虑mod3值，优先复用距离")
        else:
            if self.nr_inherit_mod30:
                print(f"规划模式: 严格mod30继承 - 只分配匹配原PCI mod30值的PCI")
            else:
                print(f"规划模式: 自由规划 - 不考虑mod30值，优先复用距离")
                
        print(f"核心规则: 只有同频同PCI的小区才需要遵循最小复用距离")
        
        # 重置统计
        self.failure_reasons = {
            'cell_not_found': [],
            'no_location': [],
            'reuse_distance_violation': [],
            'no_compliant_pci': [],
            'fallback_assignments': []
        }
        
        # 复制规划小区数据
        result_df = self.cells_to_plan.copy()
        
        # 初始化新列
        cell_names = []
        assigned_pcis = []
        original_pcis = []
        original_mods = []
        assigned_mods = []
        mod_same_flags = []
        earfcn_dls = []
        assignment_reasons = []
        predicted_min_distances = []
        
        total_cells = len(result_df)
        
        # 关键修改：按基站分组，确保同站点的小区一起处理
        # 这样可以避免同站同模问题
        print("按基站分组处理小区，确保同站点小区模值错开...")

        # 创建基站分组
        if self.network_type == "NR":
            enodeb_col = 'gNodeBID'
        else:
            enodeb_col = 'eNodeBID'

        # 关键修改：按物理位置分组处理，确保同经纬度的小区一起处理
        print("按物理位置分组处理小区，确保同站点小区模值错开...")

        # 创建位置到小区的映射
        location_groups = {}
        for idx, cell in result_df.iterrows():
            if self.network_type == "NR":
                enodeb_id = cell.get('gNodeBID')
            else:
                enodeb_id = cell.get('eNodeBID')
            cell_id = cell.get('CellID')

            # 获取小区位置信息
            cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
            if cell_info is not None:
                lat = cell_info.get('lat')
                lon = cell_info.get('lon')
                if pd.notna(lat) and pd.notna(lon):
                    # 使用经纬度作为位置键（保留6位小数精度）
                    location_key = f"{lat:.6f},{lon:.6f}"
                    if location_key not in location_groups:
                        location_groups[location_key] = []
                    location_groups[location_key].append((idx, cell))

        print(f"发现 {len(location_groups)} 个不同的物理位置")

        # 创建处理顺序：同位置的小区连续处理
        processing_order = []
        for location_key, cells in location_groups.items():
            for idx, cell in cells:
                processing_order.append((idx, cell))

        # 处理没有位置信息的小区（放在最后）
        remaining_cells = []
        for idx, cell in result_df.iterrows():
            if not any(idx == order_idx for order_idx, _ in processing_order):
                remaining_cells.append((idx, cell))
        processing_order.extend(remaining_cells)

        print(f"总共 {len(processing_order)} 个小区需要处理")

        # 按新的顺序处理小区，但保持原始索引对应关系
        # 关键修复：使用索引对应的数据结构，避免顺序错位
        results_by_index = {}

        for order_idx, (idx, cell) in enumerate(processing_order):
            # 根据网络类型选择正确的基站ID列
            if self.network_type == "NR":
                enodeb_id = cell.get('gNodeBID')
            else:
                enodeb_id = cell.get('eNodeBID')
            cell_id = cell.get('CellID')

            # 显示进度
            if (order_idx + 1) % 50 == 0 or (order_idx + 1) == len(processing_order):
                print(f"正在规划小区 {order_idx+1}/{len(processing_order)}")

            # 获取原PCI信息和小区名称
            cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
            original_pci = cell_info['pci'] if cell_info is not None else None
            # 获取小区名称并清理
            if cell_info is not None and 'cell_name' in cell_info and not pd.isna(cell_info['cell_name']):
                cell_name = str(cell_info['cell_name']).strip()
                # 过滤掉模板行的标识
                if cell_name in ['非必填', 'UserLabel', '必填']:
                    cell_name = f"小区_{enodeb_id}_{cell_id}"
            else:
                cell_name = f"小区_{enodeb_id}_{cell_id}"

            # 分配PCI（优先确保复用距离）
            assigned_pci, reason, earfcn_dl, predicted_distance = self.assign_pci_with_reuse_priority(enodeb_id, cell_id)

            # 仅当成功分配PCI时才更新小区PCI（优化：无位置信息不更新）
            if assigned_pci is not None and assigned_pci != -1:
                self.update_cell_pci(enodeb_id, cell_id, assigned_pci)

            # 记录特殊情况
            if 'not_found' in reason:
                self.failure_reasons['cell_not_found'].append(f"{enodeb_id}-{cell_id}")
            elif 'no_location' in reason:
                self.failure_reasons['no_location'].append(f"{enodeb_id}-{cell_id}")
            elif 'no_compliant' in reason:
                self.failure_reasons['no_compliant_pci'].append(f"{enodeb_id}-{cell_id}")
            elif 'fallback' in reason:
                self.failure_reasons['fallback_assignments'].append(f"{enodeb_id}-{cell_id}")

            # 计算模值
            original_mod = self.calculate_pci_mod(original_pci)
            assigned_mod = self.calculate_pci_mod(assigned_pci)

            # 比较模值
            if original_mod is not None and assigned_mod is not None:
                mod_same = '是' if original_mod == assigned_mod else '否'
            else:
                mod_same = '无法比较'

            # 关键修复：按原始索引存储结果，确保对应关系正确
            results_by_index[idx] = {
                'cell_name': cell_name,
                'original_pci': original_pci,
                'original_mod': original_mod,
                'assigned_pci': assigned_pci,
                'assigned_mod': assigned_mod,
                'mod_same': mod_same,
                'earfcn_dl': earfcn_dl,
                'assignment_reason': reason,
                'predicted_distance': predicted_distance
            }

            # 仅当成功分配PCI时才检查同站模冲突（优化：无位置信息不检查）
            if assigned_pci is not None and assigned_pci != -1:
                cell_info, _ = self.get_cell_info(enodeb_id, cell_id)
                if cell_info is not None and pd.notna(cell_info.get('lat')) and pd.notna(cell_info.get('lon')):
                    same_site_mods = set()
                    same_site_cells = self.get_same_site_cells(cell_info['lat'], cell_info['lon'], enodeb_id, cell_id)
                    for cell in same_site_cells:
                        if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                            same_site_mods.add(int(cell['pci']) % self.mod_value)

                    assigned_mod = assigned_pci % self.mod_value
                    if assigned_mod in same_site_mods:
                        print(f"      严重警告：小区{enodeb_id}-{cell_id}分配的PCI={assigned_pci}与同站点小区模{self.mod_value}冲突！")
                        print(f"      同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")
                        print(f"      冲突详情：")
                        for cell in same_site_cells:
                            if pd.notna(cell.get('pci')) and cell.get('pci') != -1:
                                cell_mod = int(cell['pci']) % self.mod_value
                                if cell_mod == assigned_mod:
                                    print(f"        - 基站{cell.get('enodeb_id', 'N/A')}-小区{cell.get('cell_id', 'N/A')}: PCI={cell['pci']}, mod{self.mod_value}={cell_mod}")
                    else:
                        print(f"      检查通过：小区{enodeb_id}-{cell_id}的PCI={assigned_pci} (mod{self.mod_value}={assigned_mod}) 无同站点冲突")
                        print(f"      同站点已有模{self.mod_value}值: {sorted(list(same_site_mods))}")
                else:
                    print(f"      警告：无法检查同站点冲突 - 缺少位置信息")
        
        # 关键修复：按原始索引正确组装数据，避免顺序错位
        # 先保存原有的列
        original_columns = list(result_df.columns)

        # 根据网络类型决定是否包含eNodeBID列
        if self.network_type == "NR":
            # NR网络：删除eNodeBID列，只保留CellID和小区名称
            print("    NR网络：删除冗余的eNodeBID列")
            filtered_columns = [col for col in original_columns if col != 'eNodeBID']
            # 在CellID后插入小区名称
            if 'CellID' in filtered_columns:
                cellid_index = filtered_columns.index('CellID')
                new_columns = filtered_columns[:cellid_index+1] + ['小区名称'] + filtered_columns[cellid_index+1:]
            else:
                new_columns = ['小区名称'] + filtered_columns
        else:
            # LTE网络：保持原有结构，在第3列插入小区名称
            new_columns = original_columns[:2] + ['小区名称'] + original_columns[2:]

        # 关键修复：按原始索引顺序组装数据，确保对应关系正确
        result_df_new = pd.DataFrame()

        # 首先组装原始列（保持原始顺序）
        for col in new_columns:
            if col == '小区名称':
                # 按原始索引顺序组装小区名称
                ordered_cell_names = []
                for idx in result_df.index:
                    if idx in results_by_index:
                        ordered_cell_names.append(results_by_index[idx]['cell_name'])
                    else:
                        # 对于没有处理结果的小区，使用默认名称
                        if self.network_type == "NR":
                            enodeb_id = result_df.loc[idx, 'gNodeBID'] if 'gNodeBID' in result_df.columns else 'Unknown'
                        else:
                            enodeb_id = result_df.loc[idx, 'eNodeBID'] if 'eNodeBID' in result_df.columns else 'Unknown'
                        cell_id = result_df.loc[idx, 'CellID'] if 'CellID' in result_df.columns else 'Unknown'
                        ordered_cell_names.append(f"小区_{enodeb_id}_{cell_id}")
                result_df_new[col] = ordered_cell_names
            elif col in original_columns:
                # 保持原始列的数据顺序
                result_df_new[col] = result_df[col].values

        # 添加其他新列（按原始索引顺序）
        ordered_original_pcis = []
        ordered_assigned_pcis = []
        ordered_original_mods = []
        ordered_assigned_mods = []
        ordered_mod_same_flags = []
        ordered_earfcn_dls = []
        ordered_assignment_reasons = []
        ordered_predicted_distances = []

        for idx in result_df.index:
            if idx in results_by_index:
                ordered_original_pcis.append(results_by_index[idx]['original_pci'])
                ordered_assigned_pcis.append(results_by_index[idx]['assigned_pci'])
                ordered_original_mods.append(results_by_index[idx]['original_mod'])
                ordered_assigned_mods.append(results_by_index[idx]['assigned_mod'])
                ordered_mod_same_flags.append(results_by_index[idx]['mod_same'])
                ordered_earfcn_dls.append(results_by_index[idx]['earfcn_dl'])
                ordered_assignment_reasons.append(results_by_index[idx]['assignment_reason'])
                ordered_predicted_distances.append(results_by_index[idx]['predicted_distance'])
            else:
                # 对于没有处理结果的小区，使用默认值
                ordered_original_pcis.append(None)
                ordered_assigned_pcis.append(None)
                ordered_original_mods.append(None)
                ordered_assigned_mods.append(None)
                ordered_mod_same_flags.append('无法比较')
                ordered_earfcn_dls.append(0.0)
                ordered_assignment_reasons.append('未处理')
                ordered_predicted_distances.append(0.0)

        result_df_new['网络类型'] = self.network_type
        result_df_new['原PCI'] = ordered_original_pcis
        result_df_new['分配的PCI'] = ordered_assigned_pcis

        # 根据网络类型设置模值列名
        if self.network_type == "LTE":
            result_df_new['原PCI模3'] = ordered_original_mods
            result_df_new['新PCI模3'] = ordered_assigned_mods
            result_df_new['模3是否相同'] = ordered_mod_same_flags
        else:
            result_df_new['原PCI模30'] = ordered_original_mods
            result_df_new['新PCI模30'] = ordered_assigned_mods
            result_df_new['模30是否相同'] = ordered_mod_same_flags

        result_df_new['earfcnDl'] = ordered_earfcn_dls
        result_df_new['分配原因'] = ordered_assignment_reasons

        # 更新result_df
        result_df = result_df_new
        
        # 计算最终验证的最小复用距离
        final_min_distances = self.calculate_final_min_reuse_distance(result_df)
        result_df['最小复用距离(km)'] = final_min_distances
        
        # 计算执行时间
        end_time = time.time()
        execution_time = end_time - start_time
        
        # 统计结果
        self.print_reuse_focused_statistics(result_df, execution_time, self.params_file)
        
        return result_df
    
    def print_reuse_focused_statistics(self, result_df: pd.DataFrame, execution_time: float, params_file: str = None):
        """
        打印复用距离优先的统计信息
        """
        print(f"\\n=== {self.network_type}网络PCI规划完成（同频PCI复用距离优先） ===")
        if params_file:
            print(f"参数文件: {os.path.basename(params_file)}")
        print(f"执行时间: {execution_time:.2f} 秒")
        print(f"平均每个小区处理时间: {execution_time/len(result_df):.4f} 秒")
        
        total_cells = len(result_df)
        success_count = len(result_df[result_df['分配的PCI'] != -1])
        
        print(f"规划统计:")
        print(f"  总小区数: {total_cells}")
        print(f"  成功分配: {success_count} ({success_count/total_cells*100:.1f}%)")
        
        # 同频PCI复用距离合规性统计
        distance_col = result_df['最小复用距离(km)']
        
        # 分类统计
        no_reuse_cells = result_df[distance_col == "无复用PCI"]
        failed_cells = result_df[distance_col == "分配失败"]
        location_missing_cells = result_df[distance_col == "位置信息缺失"]
        
        # 数值距离的小区
        numeric_distances = []
        numeric_indices = []
        for idx, dist in distance_col.items():
            if isinstance(dist, str) and dist not in ["无复用PCI", "分配失败", "位置信息缺失"]:
                try:
                    numeric_distances.append(float(dist))
                    numeric_indices.append(idx)
                except ValueError:
                    pass
        
        numeric_df = result_df.loc[numeric_indices] if numeric_indices else pd.DataFrame()
        
        if len(numeric_distances) > 0:
            compliant_numeric = [d for d in numeric_distances if d >= self.reuse_distance_km]
            violation_numeric = [d for d in numeric_distances if d < self.reuse_distance_km]
        else:
            compliant_numeric = []
            violation_numeric = []
        
        print(f"\\n同频PCI复用距离详细统计:")
        print(f"  最小复用距离要求: {self.reuse_distance_km}km (仅适用于同频同PCI小区)")
        print(f"  无复用PCI小区: {len(no_reuse_cells)} ({len(no_reuse_cells)/total_cells*100:.1f}%)")
        print(f"  有复用且合规小区: {len(compliant_numeric)} ({len(compliant_numeric)/total_cells*100:.1f}%)")
        print(f"  有复用但违规小区: {len(violation_numeric)} ({len(violation_numeric)/total_cells*100:.1f}%)")
        
        # 复用距离分布统计（仅针对数值距离）
        if numeric_distances:
            print(f"\\n复用距离分布统计 (有复用PCI的小区):")
            print(f"  平均复用距离: {np.mean(numeric_distances):.2f}km")
            print(f"  最小复用距离: {np.min(numeric_distances):.2f}km")
            print(f"  最大复用距离: {np.max(numeric_distances):.2f}km")
            print(f"  中位数复用距离: {np.median(numeric_distances):.2f}km")
            
            # 距离区间分布
            distance_ranges = [
                (0, self.reuse_distance_km, f"< {self.reuse_distance_km}km (违规)"),
                (self.reuse_distance_km, self.reuse_distance_km + 2, f"{self.reuse_distance_km}-{self.reuse_distance_km + 2}km (接近阈值)"),
                (self.reuse_distance_km + 2, self.reuse_distance_km + 5, f"{self.reuse_distance_km + 2}-{self.reuse_distance_km + 5}km (适中)"),
                (self.reuse_distance_km + 5, float('inf'), f"> {self.reuse_distance_km + 5}km (较远)")
            ]
            
            print(f"  距离区间分布:")
            for min_dist, max_dist, desc in distance_ranges:
                count = sum(1 for d in numeric_distances if min_dist <= d < max_dist)
                if count > 0:
                    print(f"    {desc}: {count} ({count/len(numeric_distances)*100:.1f}%)")
        
        if len(failed_cells) > 0:
            print(f"  分配失败小区: {len(failed_cells)} ({len(failed_cells)/total_cells*100:.1f}%)")
        if len(location_missing_cells) > 0:
            print(f"  位置信息缺失: {len(location_missing_cells)} ({len(location_missing_cells)/total_cells*100:.1f}%)")
        
        # 违规小区详情
        if len(violation_numeric) > 0:
            print(f"\\n  违规小区详情 (前10个):")
            violation_indices = [numeric_indices[i] for i, d in enumerate(numeric_distances) if d < self.reuse_distance_km][:10]
            for idx in violation_indices:
                cell = result_df.loc[idx]
                print(f"    {cell['eNodeBID']}-{cell['CellID']}: 实际距离 {cell['最小复用距离(km)']}km")
        
        # 模值继承统计
        if self.inherit_mod:
            if self.network_type == "LTE":
                mod_col = '模3是否相同'
                mod_name = "模3"
            else:
                mod_col = '模30是否相同'
                mod_name = "模30"
                
            mod_inherited = result_df[result_df[mod_col] == '是']
            mod_rate = len(mod_inherited) / total_cells * 100
            print(f"\\n{mod_name}继承统计:")
            print(f"  {mod_name}继承成功: {len(mod_inherited)} ({mod_rate:.1f}%)")
        
        # 同站点模值冲突分析
        self.analyze_same_site_mod_conflicts(result_df)
        
        # 分配原因统计
        print(f"\\n分配原因统计:")
        reason_counts = result_df['分配原因'].value_counts()
        for reason, count in reason_counts.items():
            print(f"  {reason}: {count} ({count/total_cells*100:.1f}%)")
        
        # 特殊情况统计
        special_cases = 0
        for reason, cells in self.failure_reasons.items():
            if cells:
                special_cases += len(cells)
                print(f"  {reason}: {len(cells)} 个小区")
        
        if special_cases == 0:
            print("\\n[成功] 所有小区均正常分配PCI")
        
        print(f"\\n距离状态说明:")
        print(f"  '无复用PCI': 该小区PCI在同频点内无其他小区使用，满足复用距离要求")
        print(f"  数值(如3.45): 同频同PCI小区间的最小距离(km)")
        print(f"  '分配失败': PCI分配过程失败")
        print(f"  '位置信息缺失': 小区缺少经纬度信息，无法计算距离")
        
        print(f"\\n规划完成！")
    
    def analyze_same_site_mod_conflicts(self, result_df: pd.DataFrame):
        """
        分析同站点模值冲突情况（基于经纬度判断）
        """
        print(f"\\n同位置模{self.mod_value}冲突分析:")
        
        # 按经纬度分组统计
        conflicts = []
        total_locations = 0
        conflict_locations = 0
        
        # 检查是否有经纬度列
        lat_col = '小区纬度' if '小区纬度' in result_df.columns else 'lat'
        lon_col = '小区经度' if '小区经度' in result_df.columns else 'lon'
        
        if lat_col in result_df.columns and lon_col in result_df.columns:
            # 按经纬度分组（使用经纬度组合作为键）
            location_groups = result_df.groupby([lat_col, lon_col])
            
            for (lat, lon), group in location_groups:
                total_locations += 1
                
                # 获取该位置下所有小区的PCI和模值
                location_cells = []
                for _, cell in group.iterrows():
                    pci = cell['分配的PCI']
                    if pd.notna(pci) and pci != -1:
                        mod_value = int(pci) % self.mod_value
                        # 获取基站ID用于显示
                        if self.network_type == "LTE":
                            enodeb_col = 'eNodeBID'
                        else:
                            enodeb_col = 'gNodeBID'
                        enodeb_id = cell[enodeb_col] if enodeb_col in cell else 'N/A'
                        
                        location_cells.append({
                            'cell_id': cell['CellID'],
                            'pci': int(pci),
                            'mod': mod_value,
                            'enodeb_id': enodeb_id
                        })
                
                if len(location_cells) <= 1:
                    continue  # 单小区位置无冲突可能
                
                # 检查模值冲突
                mod_values = [cell['mod'] for cell in location_cells]
                unique_mods = set(mod_values)
                
                if len(unique_mods) < len(mod_values):
                    # 发现冲突
                    conflict_locations += 1
                    
                    # 找出冲突的模值
                    conflict_mods = []
                    for mod in unique_mods:
                        mod_count = mod_values.count(mod)
                        if mod_count > 1:
                            conflict_cells = [cell for cell in location_cells if cell['mod'] == mod]
                            conflict_mods.append((mod, conflict_cells))
                    
                    conflicts.append({
                        'lat': lat,
                        'lon': lon,
                        'total_cells': len(location_cells),
                        'conflicts': conflict_mods
                    })
            
            # 输出统计结果
            print(f"  总位置数: {total_locations}")
            print(f"  有模{self.mod_value}冲突的位置: {conflict_locations} ({conflict_locations/total_locations*100:.1f}%)")
            print(f"  无模{self.mod_value}冲突的位置: {total_locations - conflict_locations} ({(total_locations - conflict_locations)/total_locations*100:.1f}%)")
            
            if conflicts:
                print(f"\\n  冲突详情 (前5个位置):")
                for i, conflict in enumerate(conflicts[:5]):
                    lat = conflict['lat']
                    lon = conflict['lon']
                    print(f"    位置({lat}, {lon}): {conflict['total_cells']}个小区")
                    for mod, cells in conflict['conflicts']:
                        cell_info = ', '.join([f"小区{cell['cell_id']}(PCI{cell['pci']})" for cell in cells])
                        print(f"      模{self.mod_value}={mod}冲突: {cell_info}")
            
            if conflict_locations == 0:
                print(f"  优秀！所有位置都避免了同位置模{self.mod_value}冲突")
            elif conflict_locations / total_locations <= 0.1:
                print(f"  良好！大部分位置避免了同位置模{self.mod_value}冲突")
            else:
                print(f"  需要关注：较多位置存在同位置模{self.mod_value}冲突")
        else:
            print(f"  警告：缺少经纬度信息，无法进行同位置模{self.mod_value}冲突分析")
        



def main():
    """
    主函数 - LTE/NR分离式PCI规划工具
    """
    print("=== LTE/NR分离式PCI规划工具 - 支持mod30逻辑 ===\\n")
    
    # 文件路径
    cells_file = "待规划小区/cell-tree-export-20250915204721.xlsx"
    
    # 动态查找全量工参文件
    params_files = glob.glob("全量工参/ProjectParameter_mongoose*.xlsx")
    if not params_files:
        print("错误: 找不到全量工参文件 (ProjectParameter_mongoose*.xlsx)")
        return

    print(f"\n发现 {len(params_files)} 个全量工参文件:")
    for i, file_path in enumerate(params_files, 1):
        print(f"  {i}. {os.path.basename(file_path)}")

    # 选择最新的工参文件
    updater = NetworkParameterUpdater()
    params_file = updater._get_latest_parameter_file(params_files)
    if not params_file:
        print("错误: 无法确定最新的全量工参文件")
        return

    print(f"\n已选择用于PCI规划的全量工参文件: {os.path.basename(params_file)}")
    print(f"完整路径: {params_file}")
    
    # 动态查找现网工参压缩包
    baseline_zips = glob.glob("全量工参/BaselineLab_*.zip")
    if not baseline_zips:
        online_params_zip = "全量工参/BaselineLab_172.16.34.163_1758250508_8172677.zip"
    else:
        online_params_zip = baseline_zips[0]
    
    # 显示主菜单
    print("请选择要执行的操作:")
    print("1. PCI规划")
    print("2. 现网工参更新")
    print("3. 退出")
    
    main_choice = input("请输入选择 (1/2/3): ").strip()
    
    if main_choice == "3":
        print("程序退出")
        return
    
    if main_choice == "2":
        # 现网工参更新功能
        print("\\n=== 现网工参更新功能 ===")
        
        # 检查压缩包是否存在
        if not os.path.exists(online_params_zip):
            print(f"错误: 找不到现网工参压缩包: {online_params_zip}")
            return
        
        # 检查全量工参文件是否存在
        if not os.path.exists(params_file):
            print(f"错误: 找不到全量工参文件: {params_file}")
            return
        
        # 执行工参更新
        updater = NetworkParameterUpdater()
        if updater.update_network_parameters():
            print("现网工参更新完成！")
        else:
            print("❌ 现网工参更新失败")
        
        # 返回到主菜单
        print("\n请选择要执行的操作:")
        print("1. PCI规划")
        print("2. 现网工参更新")
        print("3. 退出")
        
        main_choice = input("请输入选择 (1/2/3): ").strip()
        
        if main_choice == "1":
            # 继续执行PCI规划
            pass
        elif main_choice == "2":
            # 再次执行工参更新
            return main()
        elif main_choice == "3":
            print("程序退出")
            return
        else:
            print("无效选择，程序退出")
            return
    
    # 以下是原有的PCI规划功能
    # 检查两种网络类型的待规划小区
    try:
        # 检查是否存在测试文件
        test_file = "待规划小区/test_with_location.xlsx"
        if os.path.exists(test_file):
            print("检测到测试文件，使用测试模式")
            # 读取测试文件
            test_df = pd.read_excel(test_file)
            print(f"测试文件包含 {len(test_df)} 个小区")
            
            # 用户选择要规划的网络类型
            print(f"\\n请选择要规划的网络类型:")
            print(f"1. LTE ({len(test_df)} 个小区)")
            
            choice = input("请输入选择 (1): ").strip()
            
            # 用户输入最小复用距离
            while True:
                try:
                    reuse_distance = float(input(f"请输入最小PCI复用距离 (km，默认3.0): ").strip() or "3.0")
                    if reuse_distance > 0:
                        break
                    else:
                        print("复用距离必须大于0，请重新输入")
                except ValueError:
                    print("请输入有效的数字")
            
            # 根据选择执行规划
            networks_to_plan = []
            
            if choice == "1":
                networks_to_plan = ["LTE"]
            else:
                print("无效选择")
                return
            
            # 为每种网络类型执行规划
            for network_type in networks_to_plan:
                print(f"\\n{'='*50}")
                print(f"开始规划 {network_type} 网络")
                print(f"{'='*50}")
                
                # 询问继承选项
                inherit_choice = input("LTE小区是否继承原PCI的模3值？(y/n，默认n): ").strip().lower()
                lte_inherit_mod3 = inherit_choice in ['y', 'yes', '是']
                nr_inherit_mod30 = False
                
                try:
                    # 创建规划器
                    planner = LTENRPCIPlanner(
                        reuse_distance_km=reuse_distance,
                        lte_inherit_mod3=lte_inherit_mod3,
                        nr_inherit_mod30=nr_inherit_mod30,
                        network_type=network_type,
                        params_file=params_file  # 传递参数文件路径
                    )
                    
                    # 修改规划器以支持测试文件格式
                    # 临时修改cells_to_plan以适应测试文件格式
                    planner.cells_to_plan = test_df.rename(columns={
                        'eNodeBID': 'eNodeBID',
                        'CellID': 'CellID',
                        'lat': 'lat',
                        'lon': 'lon',
                        'pci': 'pci'
                    })
                    
                    # 加载参数文件
                    if network_type == "LTE":
                        params_worksheet = "LTE Project Parameters"
                    else:
                        params_worksheet = "NR Project Parameters"
                    
                    print(f"正在加载{params_worksheet}工作表...")
                    
                    # 动态查找参数文件（处理日期变化）
                    # 获取参数文件目录
                    params_dir = os.path.dirname(params_file)
                    if not params_dir:
                        params_dir = "全量工参"
                    
                    # 尝试找到正确的参数文件
                    if not os.path.exists(params_file):
                        # 查找类似文件名的文件
                        pattern = os.path.join(params_dir, "ProjectParameter_mongoose*河源电联.xlsx")
                        matching_files = glob.glob(pattern)
                        
                        if matching_files:
                            # 选择最新的文件
                            params_file = max(matching_files, key=os.path.getmtime)
                            print(f"使用替代参数文件: {params_file}")
                        else:
                            raise FileNotFoundError(f"找不到参数文件: {params_file}")
                    
                    target_raw = pd.read_excel(params_file, sheet_name=params_worksheet)
                    planner.target_cells = planner.preprocess_target_cells(target_raw)
                    planner.all_cells_combined = planner.target_cells.copy()
                    print(f"加载{network_type}小区数据: {len(planner.all_cells_combined)} 个")
                    
                    # 执行规划
                    result_df = planner.plan_pci_with_reuse_priority()
                    
                    # 生成带时间后缀的文件名
                    timestamp = planner.generate_timestamp_suffix()
                    mod_suffix = "mod3" if lte_inherit_mod3 else "nomod3"
                    output_file = f"pci_planning_{network_type.lower()}_{reuse_distance}km_{mod_suffix}_{timestamp}.xlsx"
                    
                    result_df.to_excel(output_file, index=False)
                    print(f"\\n[成功] {network_type}规划结果已保存到: {output_file}")
                    
                    # 显示前几行结果
                    print(f"\\n{network_type}规划前5行结果预览:")
                    display_columns = ['eNodeBID', 'CellID', '小区名称', '原PCI', '分配的PCI', '最小复用距离(km)', '分配原因']
                    available_columns = [col for col in display_columns if col in result_df.columns]
                    print(result_df[available_columns].head())
                    
                except Exception as e:
                    print(f"[错误] {network_type}网络规划失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            print(f"\\n[完成] 所有选定的网络类型规划完成！")
        else:
            # 原有的规划流程
            xl_file = pd.ExcelFile(cells_file)
            available_worksheets = xl_file.sheet_names
            print(f"可用的待规划小区工作表: {available_worksheets}")
            
            lte_cells_count = 0
            nr_cells_count = 0
            
            if "LTE" in available_worksheets:
                lte_df = pd.read_excel(cells_file, sheet_name="LTE")
                lte_cells_count = len(lte_df)
                print(f"LTE待规划小区数量: {lte_cells_count}")
            
            if "NR" in available_worksheets:
                nr_df = pd.read_excel(cells_file, sheet_name="NR") 
                nr_cells_count = len(nr_df)
                print(f"NR待规划小区数量: {nr_cells_count}")
            
            # 用户选择要规划的网络类型
            print(f"\\n请选择要规划的网络类型:")
            if lte_cells_count > 0:
                print(f"1. LTE ({lte_cells_count} 个小区)")
            if nr_cells_count > 0:
                print(f"2. NR ({nr_cells_count} 个小区)")
            print(f"3. 两种类型都规划")
            
            choice = input("请输入选择 (1/2/3): ").strip()
            
            # 用户输入最小复用距离
            while True:
                try:
                    reuse_distance = float(input(f"请输入最小PCI复用距离 (km，默认3.0): ").strip() or "3.0")
                    if reuse_distance > 0:
                        break
                    else:
                        print("复用距离必须大于0，请重新输入")
                except ValueError:
                    print("请输入有效的数字")
            
            # 根据选择执行规划
            networks_to_plan = []
            
            if choice == "1" and lte_cells_count > 0:
                networks_to_plan = ["LTE"]
            elif choice == "2" and nr_cells_count > 0:
                networks_to_plan = ["NR"]
            elif choice == "3":
                if lte_cells_count > 0:
                    networks_to_plan.append("LTE")
                if nr_cells_count > 0:
                    networks_to_plan.append("NR")
            else:
                print("无效选择或没有对应的待规划小区")
                return
            
            # 为每种网络类型执行规划
            for network_type in networks_to_plan:
                print(f"\\n{'='*50}")
                print(f"开始规划 {network_type} 网络")
                print(f"{'='*50}")
                
                # 根据网络类型询问继承选项
                if network_type == "LTE":
                    inherit_choice = input("LTE小区是否继承原PCI的模3值？(y/n，默认n): ").strip().lower()
                    lte_inherit_mod3 = inherit_choice in ['y', 'yes', '是']
                    nr_inherit_mod30 = False
                else:
                    inherit_choice = input("NR小区是否继承原PCI的模30值？(y/n，默认n): ").strip().lower()
                    nr_inherit_mod30 = inherit_choice in ['y', 'yes', '是']
                    lte_inherit_mod3 = False
                
                try:
                    # 创建规划器
                    planner = LTENRPCIPlanner(
                        reuse_distance_km=reuse_distance,
                        lte_inherit_mod3=lte_inherit_mod3,
                        nr_inherit_mod30=nr_inherit_mod30,
                        network_type=network_type,
                        params_file=params_file  # 传递参数文件路径
                    )

                    # 加载数据
                    planner.load_data(cells_file, params_file)
                    
                    # 执行规划
                    result_df = planner.plan_pci_with_reuse_priority()
                    
                    # 生成带时间后缀的文件名
                    timestamp = planner.generate_timestamp_suffix()
                    if network_type == "LTE":
                        mod_suffix = "mod3" if lte_inherit_mod3 else "nomod3"
                    else:
                        mod_suffix = "mod30" if nr_inherit_mod30 else "nomod30"
                        
                    output_file = f"pci_planning_{network_type.lower()}_{reuse_distance}km_{mod_suffix}_{timestamp}.xlsx"
                    
                    result_df.to_excel(output_file, index=False)
                    print(f"\\n[成功] {network_type}规划结果已保存到: {output_file}")
                    
                    # 显示前几行结果
                    print(f"\\n{network_type}规划前5行结果预览:")
                    display_columns = ['eNodeBID', 'CellID', '小区名称', '原PCI', '分配的PCI', '最小复用距离(km)', '分配原因']
                    available_columns = [col for col in display_columns if col in result_df.columns]
                    print(result_df[available_columns].head())
                    
                except Exception as e:
                    print(f"[错误] {network_type}网络规划失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            print(f"\\n[完成] 所有选定的网络类型规划完成！")
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
